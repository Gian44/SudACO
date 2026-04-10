#!/usr/bin/env python3
"""
Build results/ablation/ablation_results.xlsx from CSV inputs.

Output workbook contains exactly 2 sheets:
- Parameter tuning results
- Timeout results

Parameter sheet:
- Uses consolidated average-of-averages from consolidated_ablation_summary.csv
- Bolds exactly one row per parameter using best_config.json
- Appends a "best_config (CP-DCM-ACO)" aggregate section from best_config_results_* files,
  or from results/ablation/best_config_workbook_aggregate.json for sizes missing CSVs

Timeout sheet:
- Uses average-of-averages from timeout/*_summary.csv
- Keeps separate "Parameter Value" columns for 9x9, 16x16, 25x25 blocks
- Also adds one row per size for the ablation study wall-clock limits (5 / 20 / 120 s):
  CP-DCM-ACO uses the same aggregates as the best_config workbook section (CSV or
  best_config_workbook_aggregate.json); ACO from aggregated CP-ACS CSVs (alg 0), with
  all sizes preferring ``*_CP-ACS (100reps).csv`` when present.
"""

from __future__ import annotations

import argparse
import csv
import json
import math
from collections import defaultdict
from pathlib import Path
from statistics import mean
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SIZE_ORDER = ["9x9", "16x16", "25x25"]

# Wall-clock limits used in scripts/run_ablation.py SIZE_CONFIGS (not the timeout sweep grid).
ABLATION_WALL_TIMEOUT_S = {"9x9": 5, "16x16": 20, "25x25": 120}

# Per-size CP-ACS CSVs for default-timeout ACO rows (first existing file wins).
ACO_DEFAULT_TIMEOUT_CSV_CANDIDATES: Dict[str, List[str]] = {
    "9x9": [
        "results_9x9_CP-ACS (100reps).csv",
        "results_9x9_CP-ACS.csv",
    ],
    "16x16": [
        "results_16x16_CP-ACS (100reps).csv",
        "results_16x16_CP-ACS.csv",
    ],
    "25x25": [
        "results_25x25_CP-ACS (100reps).csv",
        "results_25x25_CP-ACS.csv",
    ],
}

PARAMETER_GROUPS = [
    ("Ant Colony System", ["nAnts", "numACS", "q0", "xi", "rho", "evap"]),
    ("Dynamic Collaborative Mechanism", ["convThresh", "entropyPct"]),
]

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def as_float(value) -> float:
    try:
        return float(value)
    except Exception:
        return float("inf")


def same_value(cell_val, target_val) -> bool:
    try:
        return abs(float(cell_val) - float(target_val)) < 1e-9
    except Exception:
        return str(cell_val).strip() == str(target_val).strip()


def read_csv_dicts(path: Path) -> List[Dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def load_param_groups(consolidated_csv: Path) -> Dict[str, List[dict]]:
    rows = read_csv_dicts(consolidated_csv)
    groups: Dict[str, List[dict]] = defaultdict(list)
    for row in rows:
        groups[row["param_name"]].append(row)
    return groups


def load_best_config_aggregates(results_root: Path) -> Dict[str, dict]:
    files = {
        "9x9": results_root / "9x9" / "best_config_results_9x9_CP-DCM-ACO.csv",
        "16x16": results_root / "16x16" / "best_config_results_16x16_CP-DCM-ACO.csv",
        "25x25": results_root / "25x25" / "best_config_results_25x25_CP-DCM-ACO.csv",
    }
    out = {}
    for size, path in files.items():
        if not path.exists():
            continue
        rows = read_csv_dicts(path)
        # Defensive dedupe by instance
        by_instance = {}
        for row in rows:
            by_instance.setdefault(row["instance"], row)
        uniq = list(by_instance.values())
        out[size] = {
            "success_mean": mean(float(x["success_%"]) for x in uniq),
            "time_mean_mean": mean(float(x["time_mean"]) for x in uniq),
            "time_std_mean": mean(float(x["time_std"]) for x in uniq),
            "cycles_mean_mean": mean(float(x["cycles_mean"]) for x in uniq),
            "instances": len(uniq),
        }

    # Fallback when CSVs were removed: results/ablation/best_config_workbook_aggregate.json
    agg_json = results_root / "ablation" / "best_config_workbook_aggregate.json"
    if agg_json.exists():
        raw = json.loads(agg_json.read_text(encoding="utf-8"))
        for size in SIZE_ORDER:
            if size in out:
                continue
            block = raw.get(size)
            if not isinstance(block, dict):
                continue
            out[size] = {
                "success_mean": float(block["success_mean"]),
                "time_mean_mean": float(block["time_mean_mean"]),
                "time_std_mean": float(block["time_std_mean"]),
                "cycles_mean_mean": float(block["cycles_mean_mean"]),
                "instances": int(block.get("instances", 0)),
            }
    return out


def load_timeout_groups(timeout_dir: Path) -> Dict[tuple, List[dict]]:
    timeout_records: List[dict] = []
    for summary_file in sorted(timeout_dir.rglob("*_summary.csv")):
        rows = read_csv_dicts(summary_file)
        if not rows:
            continue
        # Defensive dedupe by instance
        by_instance = {}
        for row in rows:
            by_instance.setdefault(row["instance"], row)
        uniq = list(by_instance.values())
        first = uniq[0]
        timeout_records.append(
            {
                "param_value": float(first["param_value"]),
                "puzzle_size": first["puzzle_size"],
                "alg": int(first["alg"]),
                "alg_name": first["alg_name"],
                "success_mean": mean(float(x["success_%"]) for x in uniq),
                "time_mean_mean": mean(float(x["time_mean"]) for x in uniq),
                "time_std_mean": mean(float(x["time_std"]) for x in uniq),
                "cycles_mean_mean": mean(float(x["cycles_mean"]) for x in uniq),
            }
        )

    grouped: Dict[tuple, List[dict]] = defaultdict(list)
    for row in timeout_records:
        grouped[(row["alg"], row["alg_name"])].append(row)
    return grouped


def load_default_ablation_timeout_cp_dcm_from_best_config(
    best_config_agg: Dict[str, dict],
) -> List[dict]:
    """
    One synthetic timeout row per puzzle size: wall time = ablation SIZE_CONFIGS timeout,
    metrics identical to the workbook's best_config (CP-DCM-ACO) aggregate block.
    """
    out: List[dict] = []
    for size in SIZE_ORDER:
        item = best_config_agg.get(size)
        if item is None:
            continue
        wall_s = float(ABLATION_WALL_TIMEOUT_S[size])
        out.append(
            {
                "param_value": wall_s,
                "puzzle_size": size,
                "alg": 2,
                "alg_name": "CP-DCM-ACO",
                "success_mean": float(item["success_mean"]),
                "time_mean_mean": float(item["time_mean_mean"]),
                "time_std_mean": float(item["time_std_mean"]),
                "cycles_mean_mean": float(item["cycles_mean_mean"]),
            }
        )
    return out


def _mean_csv_numeric(rows: List[dict], key: str) -> float:
    vals: List[float] = []
    for r in rows:
        raw = (r.get(key) or "").strip()
        if raw == "":
            continue
        try:
            vals.append(float(raw))
        except ValueError:
            continue
    return mean(vals) if vals else float("nan")


def _resolve_aco_default_timeout_csv(results_root: Path, size: str) -> Optional[Path]:
    rel = results_root / size
    for name in ACO_DEFAULT_TIMEOUT_CSV_CANDIDATES.get(size, []):
        p = rel / name
        if p.exists():
            return p
    return None


def load_default_ablation_timeout_aco(results_root: Path) -> List[dict]:
    """
    ACO (alg 0): use per-instance CP-ACS CSVs (alg 0) at ablation wall timeouts; see
    ACO_DEFAULT_TIMEOUT_CSV_CANDIDATES (prefer ``*_CP-ACS (100reps).csv`` per size).
    """
    out: List[dict] = []
    for size in SIZE_ORDER:
        path = _resolve_aco_default_timeout_csv(results_root, size)
        if path is None:
            continue
        rows = read_csv_dicts(path)
        by_instance: Dict[str, dict] = {}
        for row in rows:
            by_instance.setdefault(row["instance"], row)
        uniq = list(by_instance.values())
        uniq = [r for r in uniq if int(float(r.get("alg", -1))) == 0]
        if not uniq:
            continue
        sm = _mean_csv_numeric(uniq, "success_%")
        tmm = _mean_csv_numeric(uniq, "time_mean")
        tsm = _mean_csv_numeric(uniq, "time_std")
        cmm = _mean_csv_numeric(uniq, "cycles_mean")
        if any(math.isnan(x) for x in (sm, tmm, tsm, cmm)):
            continue
        wall_s = float(ABLATION_WALL_TIMEOUT_S[size])
        out.append(
            {
                "param_value": wall_s,
                "puzzle_size": size,
                "alg": 0,
                "alg_name": "ACO",
                "success_mean": sm,
                "time_mean_mean": tmm,
                "time_std_mean": tsm,
                "cycles_mean_mean": cmm,
            }
        )
    return out


def merge_ablation_default_timeout_rows(
    grouped: Dict[tuple, List[dict]],
    best_config_agg: Dict[str, dict],
    results_root: Path,
) -> None:
    """Add/replace ablation default-wall-timeout rows (5/20/120 s) for each algorithm."""
    bundles = [
        load_default_ablation_timeout_cp_dcm_from_best_config(best_config_agg),
        load_default_ablation_timeout_aco(results_root),
    ]
    for extras in bundles:
        if not extras:
            continue
        key = (int(extras[0]["alg"]), str(extras[0]["alg_name"]))
        existing = grouped.setdefault(key, [])
        for row in extras:
            sig_size = row["puzzle_size"]
            sig_pv = float(row["param_value"])
            replaced = False
            for i, er in enumerate(existing):
                if er["puzzle_size"] == sig_size and float(er["param_value"]) == sig_pv:
                    existing[i] = row
                    replaced = True
                    break
            if not replaced:
                existing.append(row)


def style_table_headers(ws, row_1: int, row_2: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        ws.cell(row=row_1, column=col).font = Font(bold=True)
        ws.cell(row=row_2, column=col).font = Font(bold=True)
        ws.cell(row=row_1, column=col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_2, column=col).alignment = Alignment(horizontal="center", vertical="center")


def apply_border_box(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER


def write_group_header(ws, start_row: int, title: str, max_col: int = 13) -> int:
    ws.cell(row=start_row, column=1, value=title)
    ws.cell(row=start_row, column=1).font = Font(bold=True, size=18)
    ws.cell(row=start_row, column=1).fill = PatternFill("solid", fgColor="E2F0D9")
    ws.cell(row=start_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=max_col)
    ws.row_dimensions[start_row].height = 30
    return start_row + 1


def write_standard_section(ws, start_row: int, section_title: str, rows: List[dict]) -> int:
    ws.cell(row=start_row, column=1, value=section_title)
    ws.cell(row=start_row, column=1).font = Font(bold=True)
    ws.cell(row=start_row, column=1).fill = PatternFill("solid", fgColor="D9E1F2")

    hdr1 = start_row + 1
    hdr2 = start_row + 2
    ws.cell(row=hdr1, column=1, value="Parameter Value")
    ws.merge_cells(start_row=hdr1, start_column=1, end_row=hdr2, end_column=1)

    col = 2
    for size in SIZE_ORDER:
        ws.cell(row=hdr1, column=col, value=size)
        ws.merge_cells(start_row=hdr1, start_column=col, end_row=hdr1, end_column=col + 3)
        ws.cell(row=hdr2, column=col, value="success_rate")
        ws.cell(row=hdr2, column=col + 1, value="time_mean")
        ws.cell(row=hdr2, column=col + 2, value="time_std")
        ws.cell(row=hdr2, column=col + 3, value="iter_mean")
        col += 4

    style_table_headers(ws, hdr1, hdr2, 13)

    by_param: Dict[str, Dict[str, dict]] = defaultdict(dict)
    for row in rows:
        by_param[str(row["param_value"])][row["puzzle_size"]] = row

    param_values = sorted(by_param.keys(), key=as_float)
    row_idx = hdr2 + 1
    for pv in param_values:
        ws.cell(row=row_idx, column=1, value=as_float(pv) if pv.replace(".", "", 1).isdigit() else pv)
        col = 2
        for size in SIZE_ORDER:
            item = by_param[pv].get(size)
            if item:
                ws.cell(row=row_idx, column=col, value=round(float(item["success_mean"]), 5))
                ws.cell(row=row_idx, column=col + 1, value=round(float(item["time_mean_mean"]), 5))
                ws.cell(row=row_idx, column=col + 2, value=round(float(item["time_std_mean"]), 5))
                ws.cell(row=row_idx, column=col + 3, value=round(float(item["cycles_mean_mean"]), 5))
            col += 4
        row_idx += 1

    # Border only the actual table (headers + data), not the section title row.
    apply_border_box(ws, hdr1, row_idx - 1, 1, 13)
    return row_idx + 1


def apply_best_config_bolding(ws, best_config: Dict[str, float]) -> None:
    # For each parameter section, clear data-row bold then bold exactly one row from best_config.json
    for param_name, target_value in best_config.items():
        section_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == param_name:
                section_row = r
                break
        if section_row is None:
            continue

        data_start = section_row + 3
        data_end = data_start
        while data_end <= ws.max_row and ws.cell(data_end, 1).value not in (None, ""):
            data_end += 1
        data_end -= 1

        for r in range(data_start, data_end + 1):
            for c in range(1, 14):
                ws.cell(r, c).font = Font(bold=False)

        target_row = None
        for r in range(data_start, data_end + 1):
            if same_value(ws.cell(r, 1).value, target_value):
                target_row = r
                break
        if target_row is None:
            continue

        for c in range(1, 14):
            ws.cell(target_row, c).font = Font(bold=True)


def write_best_config_section(ws, start_row: int, best_config_agg: Dict[str, dict]) -> int:
    """Metrics only: no Parameter Value column (aligns with B–M blocks above)."""
    last_metric_col = len(SIZE_ORDER) * 4
    ws.cell(row=start_row, column=1, value="best_config (CP-DCM-ACO)")
    ws.cell(row=start_row, column=1).font = Font(bold=True)
    ws.cell(row=start_row, column=1).fill = PatternFill("solid", fgColor="FCE4D6")
    ws.cell(row=start_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(
        start_row=start_row,
        start_column=1,
        end_row=start_row,
        end_column=last_metric_col,
    )

    hdr1 = start_row + 1
    hdr2 = start_row + 2
    col = 1
    for size in SIZE_ORDER:
        ws.cell(row=hdr1, column=col, value=size)
        ws.merge_cells(start_row=hdr1, start_column=col, end_row=hdr1, end_column=col + 3)
        ws.cell(row=hdr2, column=col, value="success_rate")
        ws.cell(row=hdr2, column=col + 1, value="time_mean")
        ws.cell(row=hdr2, column=col + 2, value="time_std")
        ws.cell(row=hdr2, column=col + 3, value="iter_mean")
        col += 4

    style_table_headers(ws, hdr1, hdr2, last_metric_col)

    row_idx = hdr2 + 1
    col = 1
    for size in SIZE_ORDER:
        item = best_config_agg.get(size)
        if item:
            ws.cell(row=row_idx, column=col, value=round(item["success_mean"], 5))
            ws.cell(row=row_idx, column=col + 1, value=round(item["time_mean_mean"], 5))
            ws.cell(row=row_idx, column=col + 2, value=round(item["time_std_mean"], 5))
            ws.cell(row=row_idx, column=col + 3, value=round(item["cycles_mean_mean"], 5))
        col += 4
    for c in range(1, last_metric_col + 1):
        ws.cell(row=row_idx, column=c).alignment = Alignment(horizontal="center", vertical="center")

    # Border only the table block (headers + data), not the title row.
    apply_border_box(ws, hdr1, row_idx, 1, last_metric_col)
    return row_idx + 2


def write_timeout_section_split_param(ws, start_row: int, section_title: str, rows: List[dict]) -> int:
    ws.cell(row=start_row, column=1, value=section_title)
    ws.cell(row=start_row, column=1).font = Font(bold=True)
    ws.cell(row=start_row, column=1).fill = PatternFill("solid", fgColor="D9E1F2")

    hdr1 = start_row + 1
    hdr2 = start_row + 2
    # Each size block is 5 columns:
    # [Parameter Value][success_rate][time_mean][time_std][iter_mean]
    # Size label should span only metric columns (exclude Parameter Value column).
    block_start = {"9x9": 1, "16x16": 6, "25x25": 11}

    for size in SIZE_ORDER:
        s = block_start[size]
        ws.cell(row=hdr1, column=s + 1, value=size)
        ws.merge_cells(start_row=hdr1, start_column=s + 1, end_row=hdr1, end_column=s + 4)
        ws.cell(row=hdr2, column=s, value="Parameter Value")
        ws.cell(row=hdr2, column=s + 1, value="success_rate")
        ws.cell(row=hdr2, column=s + 2, value="time_mean")
        ws.cell(row=hdr2, column=s + 3, value="time_std")
        ws.cell(row=hdr2, column=s + 4, value="iter_mean")

    style_table_headers(ws, hdr1, hdr2, 15)

    by_size = {size: [] for size in SIZE_ORDER}
    for row in rows:
        by_size[row["puzzle_size"]].append(row)
    for size in SIZE_ORDER:
        by_size[size].sort(key=lambda x: x["param_value"])

    max_len = max(len(by_size[size]) for size in SIZE_ORDER)
    row_idx = hdr2 + 1
    for i in range(max_len):
        for size in SIZE_ORDER:
            s = block_start[size]
            if i >= len(by_size[size]):
                continue
            item = by_size[size][i]
            ws.cell(row=row_idx, column=s, value=item["param_value"])
            ws.cell(row=row_idx, column=s + 1, value=round(float(item["success_mean"]), 5))
            ws.cell(row=row_idx, column=s + 2, value=round(float(item["time_mean_mean"]), 5))
            ws.cell(row=row_idx, column=s + 3, value=round(float(item["time_std_mean"]), 5))
            ws.cell(row=row_idx, column=s + 4, value=round(float(item["cycles_mean_mean"]), 5))
        row_idx += 1

    # Border only the table (headers + data), not the section title row.
    apply_border_box(ws, hdr1, row_idx - 1, 1, 15)
    return row_idx + 1


def autofit_columns(ws) -> None:
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 22)


def build_workbook(repo_root: Path, output_path: Path) -> None:
    results_root = repo_root / "results"
    ablation_root = results_root / "ablation"

    consolidated_csv = ablation_root / "consolidated_ablation_summary.csv"
    timeout_dir = ablation_root / "timeout"
    best_cfg_path = ablation_root / "best_config.json"

    param_groups = load_param_groups(consolidated_csv)
    best_config_agg = load_best_config_aggregates(results_root)
    timeout_groups = load_timeout_groups(timeout_dir)
    merge_ablation_default_timeout_rows(timeout_groups, best_config_agg, results_root)
    best_cfg = json.loads(best_cfg_path.read_text(encoding="utf-8"))

    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: parameter tuning
    ws1 = wb.create_sheet("Parameter tuning results")
    ws1["A1"] = "ABLATION - PARAMETER TUNING (AVERAGE OF AVERAGES)"
    ws1["A1"].font = Font(bold=True, size=13)
    ws1.merge_cells("A1:M1")
    ws1["A1"].alignment = Alignment(horizontal="center")

    row_ptr = 3
    assigned = set()
    for group_title, group_params in PARAMETER_GROUPS:
        params_in_group = [p for p in group_params if p in param_groups]
        if not params_in_group:
            continue
        row_ptr = write_group_header(ws1, row_ptr, group_title, max_col=13)
        for param_name in params_in_group:
            param_rows = sorted(param_groups[param_name], key=lambda x: as_float(x["param_value"]))
            row_ptr = write_standard_section(ws1, row_ptr, param_name, param_rows)
            assigned.add(param_name)

    # Fallback for any parameters not explicitly mapped into ACS/DCM groups.
    unmapped = sorted([p for p in param_groups.keys() if p not in assigned])
    if unmapped:
        row_ptr = write_group_header(ws1, row_ptr, "Other parameters", max_col=13)
        for param_name in unmapped:
            param_rows = sorted(param_groups[param_name], key=lambda x: as_float(x["param_value"]))
            row_ptr = write_standard_section(ws1, row_ptr, param_name, param_rows)

    apply_best_config_bolding(ws1, best_cfg)
    write_best_config_section(ws1, row_ptr, best_config_agg)
    autofit_columns(ws1)

    # Sheet 2: timeout
    ws2 = wb.create_sheet("Timeout results")
    ws2["A1"] = "ABLATION - TIMEOUT (AVERAGE OF AVERAGES)"
    ws2["A1"].font = Font(bold=True, size=13)
    ws2.merge_cells("A1:O1")
    ws2["A1"].alignment = Alignment(horizontal="center")

    row_ptr = 3
    for (alg, alg_name), rows in sorted(timeout_groups.items(), key=lambda x: x[0][0]):
        section = f"timeout - {alg_name} (alg={alg})"
        row_ptr = write_timeout_section_split_param(ws2, row_ptr, section, rows)
    autofit_columns(ws2)

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="Build consolidated ablation_results.xlsx workbook.")
    parser.add_argument(
        "--repo-root",
        default=".",
        help="Repository root path (default: current directory).",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Output workbook path (default: <repo>/results/ablation/ablation_results.xlsx).",
    )
    args = parser.parse_args()

    repo_root = Path(args.repo_root).resolve()
    if args.output:
        output_path = Path(args.output).resolve()
    else:
        output_path = repo_root / "results" / "ablation" / "ablation_results.xlsx"

    build_workbook(repo_root, output_path)
    print(f"Workbook written to: {output_path}")


if __name__ == "__main__":
    main()
