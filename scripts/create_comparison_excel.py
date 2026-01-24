#!/usr/bin/env python3
"""
Transform CP test results CSV into comparison Excel format.

This script reads the results CSV and creates a comparison Excel file where
each instance has side-by-side columns for CP-ACO and CP-DCM-ACO metrics.
"""

import argparse
import csv
import re
import sys
from pathlib import Path

# Check for required dependencies
try:
    import pandas as pd
except ImportError:
    print("Error: pandas is required. Install with: pip install pandas openpyxl")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


def create_comparison_excel(csv_path, output_path=None):
    """Transform CSV results into comparison Excel format."""
    csv_path = Path(csv_path)
    
    if not csv_path.exists():
        print(f"Error: CSV file not found: {csv_path}")
        return False
    
    # Read CSV
    print(f"Reading CSV file: {csv_path}")
    df = pd.read_csv(csv_path)
    
    if df.empty:
        print("Error: CSV file is empty")
        return False
    
    # Check if we have both algorithms
    algs = df['alg'].unique()
    if len(algs) < 2:
        print(f"Warning: Only found {len(algs)} algorithm(s). Both CP-ACO (0) and CP-DCM-ACO (2) are needed for comparison.")
    
    # Create comparison DataFrame
    comparison_rows = []
    
    # Group by instance and F%
    grouped = df.groupby(['instance', 'F%'])
    
    def extract_size(instance_name):
        """Extract size from instance name like 'inst25x25_40_0.txt' -> '25x25'"""
        match = re.search(r'inst(\d+x\d+)', instance_name)
        if match:
            return match.group(1)
        return ''
    
    for (instance, frac), group in grouped:
        row = {
            'instance': instance,
            'F%': frac,
        }
        
        # Get data for each algorithm
        for alg in [0, 2]:  # CP-ACO and CP-DCM-ACO
            alg_data = group[group['alg'] == alg]
            
            if not alg_data.empty:
                alg_row = alg_data.iloc[0]
                
                # Add metrics without algorithm prefix (we'll structure columns differently)
                row[f'alg{alg}_success_%'] = alg_row['success_%']
                row[f'alg{alg}_time_mean'] = alg_row['time_mean'] if pd.notna(alg_row['time_mean']) else ''
                row[f'alg{alg}_time_std'] = alg_row['time_std'] if pd.notna(alg_row['time_std']) else ''
                row[f'alg{alg}_cycles_mean'] = alg_row['cycles_mean'] if pd.notna(alg_row['cycles_mean']) else ''
            else:
                # Missing data for this algorithm
                row[f'alg{alg}_success_%'] = ''
                row[f'alg{alg}_time_mean'] = ''
                row[f'alg{alg}_time_std'] = ''
                row[f'alg{alg}_cycles_mean'] = ''
        
        comparison_rows.append(row)
    
    # Create DataFrame
    comparison_df = pd.DataFrame(comparison_rows)
    
    # Sort by F%, then by instance
    comparison_df = comparison_df.sort_values(['F%', 'instance']).reset_index(drop=True)
    
    # Determine output path
    if output_path is None:
        output_path = csv_path.parent / f'{csv_path.stem}_comparison.xlsx'
    else:
        output_path = Path(output_path)
    
    # Write to Excel with custom two-row header
    print(f"Writing comparison Excel file: {output_path}")
    
    # Create workbook and worksheet manually to have full control over headers
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Comparison'
    
    # Header formatting
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Create two-row header structure
    # Row 1: instance, CP-ACO (merged 4 cols), CP-DCM-ACO (merged 4 cols)
    # Note: We'll set CP-ACO and CP-DCM-ACO separately after merging
    # Instance column
    cell = worksheet.cell(row=1, column=1, value='instance')
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border
    
    # CP-ACO header (will be merged B1:E1)
    cell = worksheet.cell(row=1, column=2, value='CP-ACO')
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border
    
    # CP-DCM-ACO header (will be merged F1:I1)
    cell = worksheet.cell(row=1, column=6, value='CP-DCM-ACO')
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border
    
    # Row 2: empty for instance, then metric names for each algorithm
    row2 = ['', 'success rate', 'time mean', 'time std', 'cycle mean', 'success rate', 'time mean', 'time std', 'cycle mean']
    for col_idx, header_text in enumerate(row2, start=1):
        cell = worksheet.cell(row=2, column=col_idx, value=header_text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Merge cells for row 1
    # Instance column (A) spans rows 1-2
    worksheet.merge_cells('A1:A2')
    # CP-ACO (B) spans columns B-E in row 1
    worksheet.merge_cells('B1:E1')
    # CP-DCM-ACO (F) spans columns F-I in row 1
    worksheet.merge_cells('F1:I1')
    
    # Highlight color for rows where CP-ACO success rate is not 100%
    highlight_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  # Light yellow
    
    # Track instances where CP-ACO didn't achieve 100% success
    cp_aco_not_100_count = 0
    
    # Write data rows
    for row_idx, (_, data_row) in enumerate(comparison_df.iterrows(), start=3):
        # Check if CP-ACO success rate is not 100%
        cp_aco_success = data_row.get('alg0_success_%', '')
        should_highlight = False
        try:
            if cp_aco_success != '' and float(cp_aco_success) < 100.0:
                should_highlight = True
                cp_aco_not_100_count += 1
        except (ValueError, TypeError):
            pass
        
        # Instance
        cell = worksheet.cell(row=row_idx, column=1, value=data_row['instance'])
        cell.border = border
        if should_highlight:
            cell.fill = highlight_fill
        
        # CP-ACO metrics
        for col_offset, metric_key in enumerate(['alg0_success_%', 'alg0_time_mean', 'alg0_time_std', 'alg0_cycles_mean'], start=2):
            cell = worksheet.cell(row=row_idx, column=col_offset, value=data_row.get(metric_key, ''))
            cell.border = border
            if should_highlight:
                cell.fill = highlight_fill
        
        # CP-DCM-ACO metrics
        for col_offset, metric_key in enumerate(['alg2_success_%', 'alg2_time_mean', 'alg2_time_std', 'alg2_cycles_mean'], start=6):
            cell = worksheet.cell(row=row_idx, column=col_offset, value=data_row.get(metric_key, ''))
            cell.border = border
            if should_highlight:
                cell.fill = highlight_fill
        
        # Format data cells
        for col_idx in range(1, 10):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Set column widths
    column_widths = {
        'A': 25,  # instance
        'B': 12,  # CP-ACO success rate
        'C': 12,  # CP-ACO time mean
        'D': 12,  # CP-ACO time std
        'E': 12,  # CP-ACO cycles mean
        'F': 12,  # CP-DCM-ACO success rate
        'G': 12,  # CP-DCM-ACO time mean
        'H': 12,  # CP-DCM-ACO time std
        'I': 12,  # CP-DCM-ACO cycles mean
    }
    
    for col, width in column_widths.items():
        worksheet.column_dimensions[col].width = width
    
    # Add summary row at the bottom
    summary_row = len(comparison_df) + 3
    summary_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')  # Light gray
    summary_font = Font(bold=True, size=11)
    
    # Summary label
    cell = worksheet.cell(row=summary_row, column=1, value=f'Instances where CP-ACO did not achieve 100% success rate:')
    cell.fill = summary_fill
    cell.font = summary_font
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.border = border
    
    # Summary count
    cell = worksheet.cell(row=summary_row, column=2, value=cp_aco_not_100_count)
    cell.fill = summary_fill
    cell.font = summary_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border
    
    # Merge summary cells across all columns for better appearance
    for col_idx in range(3, 10):
        cell = worksheet.cell(row=summary_row, column=col_idx)
        cell.fill = summary_fill
        cell.border = border
    
    # Freeze header rows
    worksheet.freeze_panes = 'A3'
    
    # Save workbook
    workbook.save(output_path)
    
    print(f"✓ Successfully created comparison Excel file: {output_path}")
    print(f"  Total instances compared: {len(comparison_df)}")
    
    return True


def main():
    ap = argparse.ArgumentParser(
        description='Transform CP test results CSV into comparison Excel format with side-by-side algorithm metrics.'
    )
    ap.add_argument('csv_file', help='Input CSV file path')
    ap.add_argument('-o', '--output', help='Output Excel file path (default: <csv_name>_comparison.xlsx)')
    args = ap.parse_args()
    
    success = create_comparison_excel(args.csv_file, args.output)
    sys.exit(0 if success else 1)


if __name__ == '__main__':
    main()
