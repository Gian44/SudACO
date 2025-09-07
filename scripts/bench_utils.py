#!/usr/bin/env python3
"""Shared benchmarking utilities for Sudoku ACO experiments.

This module provides helpers for running the solver over the logic-solvable
and general instance sets.  It is intended to be imported by small wrapper
scripts that handle the command line interface for each benchmark type.
"""

import csv
import math
import os
import re
import subprocess
from pathlib import Path
from statistics import mean, pstdev


def default_binary():
    """Guess a sensible default solver binary depending on the platform."""
    if os.name == 'nt':
        cand = Path('vs2017') / 'x64' / 'Release' / 'sudoku_ants.exe'
        if cand.exists():
            return str(cand)
        cand = Path('vs2017') / 'Release' / 'sudoku_ants.exe'
        if cand.exists():
            return str(cand)
        return 'sudoku_ants.exe'
    else:
        return './sudokusolver'


def run_solver(binary, file_path, alg, timeout, extra_args=None):
    """Invoke the solver and parse its output."""
    args = [binary, '--file', str(file_path), '--alg', str(alg), '--timeout', str(timeout)]
    if extra_args:
        args.extend(extra_args)
    try:
        out = subprocess.check_output(args, stderr=subprocess.STDOUT, universal_newlines=True)
    except subprocess.CalledProcessError as e:
        # Non-zero exit; try to parse any output
        out = e.output

    # Parse output: expected two lines at the end: success_flag (0=success), time
    # Be robust to extra prints and scientific notation times like 5e-05.
    lines = [ln.strip() for ln in out.splitlines() if ln.strip() != '']
    success = False
    elapsed = math.nan
    cycles = math.nan

    # Prefer an explicit success flag line equal to '0' or '1'
    flag = None
    for ln in lines:
        if ln in ('0', '1'):
            flag = ln

    # Robust float pattern supporting scientific notation
    float_pat = re.compile(r"[-+]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][-+]?\d+)?")
    # Try to find the last float in the output (time line or 'solved in X')
    for ln in reversed(lines):
        m = float_pat.search(ln)
        if m:
            try:
                elapsed = float(m.group(0))
                break
            except ValueError:
                continue

    # Extract number of cycles if present (supports single- and multi-colony prints)
    cyc_pat = re.compile(r"Number of cycles(?: \(multi\))?:\s*(\d+)")
    for ln in lines:
        m = cyc_pat.search(ln)
        if m:
            try:
                cycles = int(m.group(1))
            except ValueError:
                pass

    if flag is not None:
        success = (flag == '0')
    else:
        # Fallback: if we saw "solved in" treat as success
        for ln in lines:
            if 'solved in' in ln.lower():
                success = True
                break

    return success, elapsed, cycles, out


def scan_logic_instances(base):
    """Return a sorted list of logic-solvable instance files."""
    p = Path(base)
    files = sorted(p.glob('*.txt'))
    return files


def parse_general_filename(name):
    """Parse a general instance filename, returning (size, F%)."""
    stem = Path(name).stem
    parts = stem.split('_')
    if len(parts) < 3:
        return None, None
    size = parts[0].replace('inst', '')
    try:
        frac = int(parts[1])
    except ValueError:
        frac = None
    return size, frac


def scan_general_groups(base):
    """Group general instances by size and fraction of given numbers."""
    groups = {}  # (size, frac) -> [files]
    for fp in sorted(Path(base).glob('*.txt')):
        size, frac = parse_general_filename(fp.name)
        if size is None or frac is None:
            continue
        groups.setdefault((size, frac), []).append(fp)
    return groups


def safe_mean(vals):
    vals = [v for v in vals if not math.isnan(v)]
    if not vals:
        return math.nan
    return mean(vals)


def safe_std(vals):
    vals = [v for v in vals if not math.isnan(v)]
    if len(vals) < 2:
        return 0.0 if len(vals) == 1 else math.nan
    return pstdev(vals)


def write_csv(path, headers, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, 'w', newline='') as f:
        w = csv.writer(f)
        w.writerow(headers)
        for r in rows:
            w.writerow(r)


def maybe_write_xlsx(path, sheets):
    """Attempt to write an Excel workbook with the given sheets."""
    # sheets: list of (sheet_name, headers, rows)
    try:
        import openpyxl  # type: ignore
        from openpyxl import Workbook
    except Exception:
        try:
            import xlsxwriter  # type: ignore
        except Exception:
            return False
        # xlsxwriter path
        path.parent.mkdir(parents=True, exist_ok=True)
        wb = xlsxwriter.Workbook(str(path))
        for name, headers, rows in sheets:
            ws = wb.add_worksheet(name[:31])
            for j, h in enumerate(headers):
                ws.write(0, j, h)
            for i, row in enumerate(rows, start=1):
                for j, val in enumerate(row):
                    ws.write(i, j, val)
        wb.close()
        return True
    # openpyxl path
    wb = Workbook()
    # remove default sheet
    wb.remove(wb.active)
    for name, headers, rows in sheets:
        ws = wb.create_sheet(name[:31])
        ws.append(headers)
        for r in rows:
            ws.append(r)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    return True


def run_logic(algs, logic_dir, binary, timeout, reps_logic, vlog):
    """Run benchmarks on logic-solvable instances."""
    logic_rows = []
    logic_headers = ['alg', 'instance', 'success_%', 'time_mean', 'time_std', 'cycles_mean']
    logic_files = scan_logic_instances(logic_dir)
    for alg in algs:
        for idx, fp in enumerate(logic_files, start=1):
            vlog(f"[logic-solvable] alg={alg} instance {idx}/{len(logic_files)}: {fp.name}")
            successes = 0
            times = []
            cycles_solved = []
            for r in range(reps_logic):
                vlog(f"  - rep {r+1}/{reps_logic}")
                ok, t, cyc, _ = run_solver(binary, fp, alg, timeout)
                if ok:
                    successes += 1
                    times.append(t)
                    if not math.isnan(cyc):
                        cycles_solved.append(cyc)
            succ_pct = (successes / float(reps_logic)) * 100.0
            cycles_mean_val = round(safe_mean(cycles_solved), 3)
            logic_rows.append([
                alg,
                fp.name,
                round(succ_pct, 2),
                round(safe_mean(times), 6),
                round(safe_std(times), 6),
                cycles_mean_val,
            ])
            vlog(
                f"  => success%={round(succ_pct,2)} time_mean={round(safe_mean(times),6)} "
                f"time_std={round(safe_std(times),6)} cycles_mean={cycles_mean_val}"
            )
    return logic_headers, logic_rows


def run_general(algs, gen_dir, binary, timeout, vlog):
    """Run benchmarks on general instances."""
    gen_rows = []
    gen_headers = ['alg', 'puzzle', 'F%', 'solution_rate', 'time_mean', 'time_std', 'cycles_mean']
    groups = scan_general_groups(gen_dir)
    for alg in algs:
        for (size, frac), files in sorted(groups.items()):
            vlog(f"[general] alg={alg} size={size} F%={frac} files={len(files)}")
            solved = 0
            times = []
            cycles_solved = []
            for i, fp in enumerate(files, start=1):
                vlog(f"  - file {i}/{len(files)}: {fp.name}")
                ok, t, cyc, _ = run_solver(binary, fp, alg, timeout)
                if ok:
                    solved += 1
                    times.append(t)
                    if not math.isnan(cyc):
                        cycles_solved.append(cyc)
            rate = (solved / float(len(files))) * 100.0 if files else 0.0
            gen_rows.append([
                alg,
                size,
                frac,
                round(rate, 2),
                round(safe_mean(times), 6),
                round(safe_std(times), 6),
                round(safe_mean(cycles_solved), 3),
            ])
            vlog(
                f"  => solution_rate={round(rate,2)} time_mean={round(safe_mean(times),6)} "
                f"time_std={round(safe_std(times),6)} cycles_mean={round(safe_mean(cycles_solved),3)}"
            )
    return gen_headers, gen_rows

