#!/usr/bin/env python3
"""
Compare ACO (algorithm 0) vs CP-DCM-ACO (algorithm 2) at multiple wall-clock timeouts.

For **CP-DCM-ACO (alg 2)**, hyperparameters come from ``results/ablation/best_config.json``
(``run_ablation.py --consolidate``). **ACO (alg 0)** uses no JSON overrides: the binary’s
built-in defaults (``solvermain.cpp`` / CLI defaults) apply; DCM-only flags are not passed.

Outputs (default ``results/ablation/timeout``):
  alg_<id>/<timeout>_<size>_summary.csv — same row schema as ablation summaries
  timeout_comparison.xlsx — aggregated sheet per puzzle size (after consolidation, in same folder)
  results/ablation/ablation_results.xlsx — updated with timeout tables:
    - ACO table
    - CP-DCM-ACO table
    - consolidated timeout summary (CSV-like columns)

Parallel mode:
  --workers-per-alg 4
    Per puzzle size, each worker runs the full timeout grid for its algorithm.
    Start 4 workers on ACO (alg 0) and 4 on CP-DCM-ACO (alg 2).
    Workers use atomic per-instance claims (safe work-stealing), and once one
    algorithm finishes, those worker slots can transfer to help the other
    algorithm finish the same size phase.

Logging (default ``logs/timeout_comparison/``):
  timeout_orchestrator.log — parent process: phases, worker launches, Excel consolidation
  timeout_alg0.log — ACO (merged from parallel worker sessions, or direct tee in serial mode)
  timeout_alg2.log — CP-DCM-ACO (same)

  pip install openpyxl
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import os
import shutil
import subprocess
import sys
import time
from collections import OrderedDict
from datetime import datetime
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from bench_utils import (  # noqa: E402
    SolverInterruptedError,
    default_binary,
    run_solver,
    safe_mean,
    safe_std,
)

from scripts.run_ablation import (  # noqa: E402
    DEFAULTS,
    PROGRESS_HEADERS,
    SIZE_CONFIGS,
    SUMMARY_HEADERS,
    append_csv_row,
    build_solver_args_from_full_config,
    delete_ablation_progress_if_summary_done,
    ensure_csv_header,
    format_param_value,
    read_completed_from_summary,
    read_progress,
    scan_instances,
    sort_summary_csv_if_complete,
)

DEFAULT_OUTDIR = Path('results') / 'ablation' / 'timeout'
DEFAULT_BEST_CONFIG = Path('results') / 'ablation' / 'best_config.json'
DEFAULT_EXCEL = DEFAULT_OUTDIR / 'timeout_comparison.xlsx'
DEFAULT_ABLATION_EXCEL = Path('results') / 'ablation' / 'ablation_results.xlsx'
DEFAULT_LOG_DIR = REPO_ROOT / 'logs' / 'timeout_comparison'

# Wall-clock timeouts (seconds) to compare; same grid as the former ablation timeout sweep.
TIMEOUTS_PER_SIZE = OrderedDict([
    ('9x9', [1, 3, 7, 9]),
    ('16x16', [10, 15, 25, 30]),
    ('25x25', [60, 90, 150, 180]),
])

ALGORITHMS = (
    (0, 'ACO'),
    (2, 'CP-DCM-ACO'),
)

SIZE_SHORT = {'9x9': '9', '16x16': '16', '25x25': '25'}
SIZE_SORT = {'9x9': 0, '16x16': 1, '25x25': 2}


def _log_timestamp() -> str:
    return datetime.now().strftime('%Y-%m-%d %H:%M:%S')


class _TeeStdout:
    """Duplicate stdout to a log file (and optionally the real console)."""

    def __init__(self, primary, secondary):
        self._primary = primary
        self._secondary = secondary

    def write(self, data: str) -> int:
        self._primary.write(data)
        self._secondary.write(data)
        return len(data)

    def flush(self) -> None:
        self._primary.flush()
        self._secondary.flush()

    def isatty(self) -> bool:
        return self._primary.isatty()


def _append_session_to_alg_log(
    log_dir: Path,
    alg: int,
    session_path: Path,
    header: str,
) -> None:
    """Append one worker session file into the per-algorithm log, then remove the session."""
    dest = log_dir / f'timeout_alg{alg}.log'
    with open(dest, 'a', encoding='utf-8', newline='\n') as out:
        out.write('\n')
        out.write('=' * 72 + '\n')
        out.write(f'[{_log_timestamp()}] {header}\n')
        out.write('=' * 72 + '\n')
        if session_path.exists():
            with open(session_path, encoding='utf-8', errors='replace') as inc:
                shutil.copyfileobj(inc, out)
    try:
        session_path.unlink()
    except OSError:
        pass


def _try_claim_instance(
    claim_dir: Path,
    instance_name: str,
    *,
    worker_id: int,
    num_workers: int,
    stale_seconds: int = 12 * 60 * 60,
) -> Path | None:
    """
    Atomically claim one instance using O_EXCL file creation.
    Returns claim path on success, else None.
    """
    claim_dir.mkdir(parents=True, exist_ok=True)
    claim_path = claim_dir / f'{instance_name}.claim'

    def _create() -> bool:
        try:
            fd = os.open(str(claim_path), os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            try:
                msg = (
                    f'worker_id={worker_id}\n'
                    f'num_workers={num_workers}\n'
                    f'claimed_at={time.time():.6f}\n'
                )
                os.write(fd, msg.encode('utf-8', errors='replace'))
            finally:
                os.close(fd)
            return True
        except FileExistsError:
            return False
        except OSError:
            return False

    if _create():
        return claim_path

    # Reclaim stale claims from crashed workers.
    try:
        age = time.time() - claim_path.stat().st_mtime
        if age > float(stale_seconds):
            try:
                claim_path.unlink()
            except OSError:
                return None
            if _create():
                return claim_path
    except OSError:
        return None

    return None


def _release_claim(claim_path: Path | None) -> None:
    if claim_path is None:
        return
    try:
        claim_path.unlink()
    except OSError:
        pass


def load_best_config(path: Path) -> dict:
    """
    Return ``{ '9x9': cfg, '16x16': cfg, '25x25': cfg }`` with the same hyperparameters
    for every size (from a flat ``best_config.json``). Legacy per-size JSON is still
    accepted if top-level keys are puzzle sizes.
    """
    if not path.exists():
        raise SystemExit(
            f'Missing {path}. Run: python scripts/run_ablation.py --consolidate'
        )
    with open(path, 'r', encoding='utf-8') as f:
        raw = json.load(f)

    def merged_cfg(update: dict) -> dict:
        m = dict(DEFAULTS)
        m.update(update)
        return m

    # Legacy: {"9x9": {...}, "16x16": {...}, ...}
    if any(k in SIZE_CONFIGS for k in raw):
        out = {}
        for size, cfg in raw.items():
            if size not in SIZE_CONFIGS:
                continue
            out[size] = merged_cfg(cfg if isinstance(cfg, dict) else {})
        for size in SIZE_CONFIGS:
            if size not in out:
                out[size] = dict(DEFAULTS)
        return out

    # Current: flat {"nAnts": 5, ...}
    single = merged_cfg(raw)
    return {sz: dict(single) for sz in SIZE_CONFIGS}


def run_timeout_job(
    binary,
    alg: int,
    alg_name: str,
    timeout_sec: int,
    size_name: str,
    size_cfg: dict,
    extra_args: list,
    reps: int,
    outdir: Path,
    vlog,
    worker_id: int = 0,
    num_workers: int = 1,
    dynamic_claims: bool = False,
):
    """One (algorithm, timeout, size) matrix; param_value column stores timeout for traceability."""
    val_str = format_param_value('timeout', timeout_sec)
    tag = f'alg{alg} timeout={val_str}s [{size_name}]'

    sub = outdir / f'alg_{alg}'
    sub.mkdir(parents=True, exist_ok=True)

    progress_file = sub / f'{val_str}_{size_name}_progress.csv'
    summary_file = sub / f'{val_str}_{size_name}_summary.csv'
    timeout = int(timeout_sec)

    instances_all = scan_instances(size_cfg['dir'])
    if not instances_all:
        vlog(f'  No instances in {size_cfg["dir"]}')
        return

    all_instance_names = {fp.name for fp in instances_all}
    num_workers = max(1, int(num_workers))
    worker_id = int(worker_id) % num_workers
    dynamic_claims = bool(dynamic_claims and num_workers > 1)

    if dynamic_claims:
        instances = instances_all
    else:
        instances = [
            fp for i, fp in enumerate(instances_all)
            if (i % num_workers) == worker_id
        ]
        if not instances:
            vlog(f'  Worker {worker_id}/{num_workers} has no instances')
            return

    completed = read_completed_from_summary(summary_file)
    completed_overall = len(completed.intersection(all_instance_names))
    if completed_overall == len(instances_all):
        sort_summary_csv_if_complete(summary_file, [fp.name for fp in instances_all])
        delete_ablation_progress_if_summary_done(
            progress_file, summary_file, all_instance_names, vlog, tag)
        vlog(f'  [{tag}] Already complete. Skipping.')
        return

    progress = read_progress(progress_file)
    ensure_csv_header(summary_file, SUMMARY_HEADERS)
    ensure_csv_header(progress_file, PROGRESS_HEADERS)

    subset_names = {fp.name for fp in instances}
    completed_in_subset = completed.intersection(subset_names)
    total = len(instances)
    if completed_in_subset:
        mode = 'claim' if dynamic_claims else 'shard'
        vlog(f'  [{tag}] Worker {worker_id}/{num_workers} ({mode}) resuming '
             f'{len(completed_in_subset)}/{total} instances')

    claim_dir = sub / '.claims' / f'{val_str}_{size_name}'

    def _process_instance(fp: Path, idx_label: str):
        rep_map = progress.get(fp.name, {})
        done_reps = set(rep_map.keys())

        successes = 0
        times = []
        cycles_solved = []
        for _r, (succ, t, cyc) in sorted(rep_map.items()):
            if succ:
                successes += 1
                times.append(t)
                if not math.isnan(cyc):
                    cycles_solved.append(cyc)

        status = f'RESUME {len(done_reps)}/{reps}' if done_reps else ''
        vlog(f'  [{tag}] {idx_label} {fp.name} {status}')

        for rep in range(1, reps + 1):
            if rep in done_reps:
                continue
            try:
                success, t, cyc, out = run_solver(
                    binary, fp, alg, timeout, extra_args=extra_args)
            except SolverInterruptedError:
                vlog(f'    Rep {rep}/{reps}: INTERRUPTED; not recorded (will resume)')
                break

            status_str = 'OK' if success else 'FAIL'
            t_str = f'{t:.4f}s' if not math.isnan(t) else 'N/A'
            vlog(f'    Rep {rep}/{reps}: {status_str} (time={t_str})')

            prog_row = [fp.name, rep, 1 if success else 0,
                        '' if math.isnan(t) else t,
                        '' if math.isnan(cyc) else cyc]
            append_csv_row(progress_file, prog_row)

            rep_map[rep] = (success, t, cyc)
            done_reps.add(rep)
            if success:
                successes += 1
                times.append(t)
                if not math.isnan(cyc):
                    cycles_solved.append(cyc)

        if len(done_reps) < reps:
            vlog(f'    => partial ({len(done_reps)}/{reps})')
            progress[fp.name] = rep_map
            return

        succ_pct = (successes / float(reps)) * 100.0
        tm = safe_mean(times)
        ts = safe_std(times)
        cm = safe_mean(cycles_solved)
        cs = safe_std(cycles_solved)

        row = [
            timeout_sec, size_name, fp.name,
            alg, alg_name,
            round(succ_pct, 2),
            round(tm, 6) if not math.isnan(tm) else '',
            round(ts, 6) if not math.isnan(ts) else '',
            round(cm, 3) if not math.isnan(cm) else '',
            round(cs, 3) if not math.isnan(cs) else '',
        ]
        if append_csv_row(summary_file, row):
            completed.add(fp.name)
            completed_now = read_completed_from_summary(summary_file)
            if all_instance_names <= completed_now:
                sort_summary_csv_if_complete(summary_file, [p.name for p in instances_all])
                delete_ablation_progress_if_summary_done(
                    progress_file, summary_file, all_instance_names, vlog, tag)
        else:
            vlog(f'    ERROR: could not write summary for {fp.name}')
        progress[fp.name] = rep_map

    if dynamic_claims:
        idle_loops = 0
        while True:
            completed_now = read_completed_from_summary(summary_file)
            done_global = len(completed_now.intersection(all_instance_names))
            if done_global >= len(instances_all):
                break

            claimed_fp = None
            claimed_path = None
            for idx, fp in enumerate(instances_all, 1):
                if fp.name in completed_now:
                    continue
                cp = _try_claim_instance(
                    claim_dir,
                    fp.name,
                    worker_id=worker_id,
                    num_workers=num_workers,
                )
                if cp is not None:
                    claimed_fp = fp
                    claimed_path = cp
                    try:
                        _process_instance(fp, f'({idx}/{len(instances_all)}) [claim]')
                    finally:
                        _release_claim(claimed_path)
                    idle_loops = 0
                    break

            if claimed_fp is None:
                idle_loops += 1
                if idle_loops == 1 or idle_loops % 20 == 0:
                    vlog(
                        f'  [{tag}] Worker {worker_id}/{num_workers} waiting for claimable '
                        f'instances ({done_global}/{len(instances_all)} complete)'
                    )
                time.sleep(0.5)
    else:
        for idx, fp in enumerate(instances, 1):
            if fp.name in completed:
                continue
            _process_instance(fp, f'({idx}/{total})')

    delete_ablation_progress_if_summary_done(
        progress_file, summary_file, all_instance_names, vlog, tag)


def wait_for_timeout_matrix_complete(
    *,
    outdir: Path,
    alg: int,
    timeout_sec: int,
    size_name: str,
    size_cfg: dict,
    alg_name: str,
    vlog,
    poll_seconds: float = 2.0,
) -> None:
    """Block until summary for (alg, timeout, size) contains every instance."""
    instances_all = scan_instances(size_cfg['dir'])
    all_instance_names = {fp.name for fp in instances_all}
    total = len(all_instance_names)
    if total == 0:
        return

    val_str = format_param_value('timeout', timeout_sec)
    summary_file = outdir / f'alg_{alg}' / f'{val_str}_{size_name}_summary.csv'

    last_done = -1
    last_log_ts = 0.0
    while True:
        completed = read_completed_from_summary(summary_file)
        done = len(completed.intersection(all_instance_names))
        if done >= total:
            if last_done != done:
                vlog(
                    f'  [BARRIER] {alg_name} timeout={timeout_sec}s {size_name}: '
                    f'global complete ({done}/{total}); proceeding.'
                )
            return

        now = time.time()
        if done != last_done or (now - last_log_ts) >= 30.0:
            vlog(
                f'  [BARRIER] {alg_name} timeout={timeout_sec}s {size_name}: '
                f'waiting for global completion ({done}/{total})'
            )
            last_done = done
            last_log_ts = now
        time.sleep(max(0.2, float(poll_seconds)))


def collect_timeout_aggregates(outdir: Path):
    """
    Aggregate summary CSVs into rows per (timeout, puzzle_size, algorithm).
    """
    rows_agg = []
    for alg_id, alg_name in ALGORITHMS:
        sub = outdir / f'alg_{alg_id}'
        if not sub.exists():
            continue
        for csv_file in sorted(sub.glob('*_summary.csv')):
            base = csv_file.name[:-len('_summary.csv')]
            try:
                timeout_part_s, size_name = base.split('_', 1)
                timeout_part = int(float(timeout_part_s))
            except Exception:
                continue
            if size_name not in SIZE_CONFIGS:
                continue

            by_instance = {}
            try:
                with open(csv_file, 'r', newline='') as f:
                    for row in csv.DictReader(f):
                        inst = (row.get('instance') or '').strip()
                        if inst:
                            # Defensive de-dup: keep last row per instance.
                            by_instance[inst] = row
            except Exception:
                continue

            if not by_instance:
                continue

            succ_vals = []
            tmean_vals = []
            tstd_vals = []
            cmean_vals = []
            cstd_vals = []
            for row in by_instance.values():
                try:
                    succ_vals.append(float(row.get('success_%') or 'nan'))
                except Exception:
                    succ_vals.append(float('nan'))
                try:
                    tmean_vals.append(float(row.get('time_mean') or 'nan'))
                except Exception:
                    tmean_vals.append(float('nan'))
                try:
                    tstd_vals.append(float(row.get('time_std') or 'nan'))
                except Exception:
                    tstd_vals.append(float('nan'))
                try:
                    cmean_vals.append(float(row.get('cycles_mean') or 'nan'))
                except Exception:
                    cmean_vals.append(float('nan'))
                try:
                    cstd_vals.append(float(row.get('cycles_std') or 'nan'))
                except Exception:
                    cstd_vals.append(float('nan'))

            rows_agg.append({
                'timeout_s': timeout_part,
                'puzzle_size': size_name,
                'alg': alg_id,
                'alg_name': alg_name,
                'instances': len(succ_vals),
                'success_mean': safe_mean(succ_vals),
                'time_mean_mean': safe_mean(tmean_vals),
                'time_std_mean': safe_mean(tstd_vals),
                'cycles_mean_mean': safe_mean(cmean_vals),
                'cycles_std_mean': safe_mean(cstd_vals),
            })

    rows_agg.sort(key=lambda r: (SIZE_SORT.get(r['puzzle_size'], 99), r['timeout_s'], r['alg']))
    return rows_agg


def _write_sheet_table(ws, row_idx, title, headers, rows, header_font, header_fill, thin_border):
    ws.cell(row=row_idx, column=1, value=title).font = header_font
    row_idx += 1
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    row_idx += 1
    for r in rows:
        for c, v in enumerate(r, 1):
            cell = ws.cell(row=row_idx, column=c, value=v)
            cell.border = thin_border
        row_idx += 1
    return row_idx + 1


def consolidate_timeout_excel(outdir: Path, excel_path: Path) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side
    except ImportError:
        print('ERROR: openpyxl required. pip install openpyxl')
        return False

    rows_agg = collect_timeout_aggregates(outdir)
    if not rows_agg:
        print('No timeout comparison summaries found to consolidate.')
        return False

    wb = Workbook()
    wb.remove(wb.active)
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    for size_name in SIZE_CONFIGS:
        size_rows = [r for r in rows_agg if r['puzzle_size'] == size_name]
        if not size_rows:
            continue
        title = f'Timeout cmp {size_name}'[:31]
        ws = wb.create_sheet(title=title)
        headers = [
            'Timeout (s)', 'Algorithm ID', 'Algorithm',
            'Instances', 'Mean success %', 'Mean time (s)',
        ]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        for r_idx, r in enumerate(size_rows, 2):
            vals = [
                r['timeout_s'], r['alg'], r['alg_name'],
                r['instances'],
                round(r['success_mean'], 5) if not math.isnan(r['success_mean']) else '',
                round(r['time_mean_mean'], 5) if not math.isnan(r['time_mean_mean']) else '',
            ]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=r_idx, column=c, value=v)
                cell.border = thin_border
        ws.freeze_panes = 'A2'

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(excel_path))
    print(f'Saved: {excel_path}')
    return True


def merge_timeout_into_ablation_workbook(outdir: Path, ablation_excel_path: Path) -> bool:
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill, Border, Side
    except ImportError:
        print('ERROR: openpyxl required. pip install openpyxl')
        return False

    rows_agg = collect_timeout_aggregates(outdir)
    if not rows_agg:
        print('No timeout comparison summaries found to merge into ablation workbook.')
        return False

    if ablation_excel_path.exists():
        wb = load_workbook(str(ablation_excel_path))
    else:
        wb = Workbook()

    if 'Timeout results' in wb.sheetnames:
        del wb['Timeout results']
    if 'Sheet' in wb.sheetnames and len(wb.sheetnames) == 1:
        ws0 = wb['Sheet']
        if ws0.max_row <= 1 and ws0.max_column <= 1 and ws0['A1'].value is None:
            del wb['Sheet']

    ws = wb.create_sheet(title='Timeout results')
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    aco_rows = []
    dcm_rows = []
    for r in rows_agg:
        out_row = [
            r['puzzle_size'], r['timeout_s'], r['instances'],
            round(r['success_mean'], 5) if not math.isnan(r['success_mean']) else '',
            round(r['time_mean_mean'], 5) if not math.isnan(r['time_mean_mean']) else '',
            round(r['time_std_mean'], 5) if not math.isnan(r['time_std_mean']) else '',
            round(r['cycles_mean_mean'], 5) if not math.isnan(r['cycles_mean_mean']) else '',
            round(r['cycles_std_mean'], 5) if not math.isnan(r['cycles_std_mean']) else '',
        ]
        if r['alg'] == 0:
            aco_rows.append(out_row)
        elif r['alg'] == 2:
            dcm_rows.append(out_row)

    table_headers = [
        'Puzzle size', 'Timeout (s)', 'Instances',
        'Mean success %', 'Mean time (s)', 'Mean time std (s)',
        'Mean cycles', 'Mean cycles std',
    ]
    row_idx = 1
    row_idx = _write_sheet_table(
        ws, row_idx, 'ACO timeout results', table_headers, aco_rows,
        header_font, header_fill, thin_border)
    row_idx = _write_sheet_table(
        ws, row_idx, 'CP-DCM-ACO timeout results', table_headers, dcm_rows,
        header_font, header_fill, thin_border)

    consolidated_headers = [
        'param_name', 'param_value', 'puzzle_size', 'alg', 'alg_name',
        'instances', 'success_mean', 'time_mean_mean', 'time_std_mean',
        'cycles_mean_mean', 'cycles_std_mean',
    ]
    consolidated_rows = []
    for r in rows_agg:
        consolidated_rows.append([
            'timeout',
            r['timeout_s'],
            r['puzzle_size'],
            r['alg'],
            r['alg_name'],
            r['instances'],
            round(r['success_mean'], 5) if not math.isnan(r['success_mean']) else '',
            round(r['time_mean_mean'], 5) if not math.isnan(r['time_mean_mean']) else '',
            round(r['time_std_mean'], 5) if not math.isnan(r['time_std_mean']) else '',
            round(r['cycles_mean_mean'], 5) if not math.isnan(r['cycles_mean_mean']) else '',
            round(r['cycles_std_mean'], 5) if not math.isnan(r['cycles_std_mean']) else '',
        ])
    _write_sheet_table(
        ws, row_idx, 'Consolidated timeout summary', consolidated_headers, consolidated_rows,
        header_font, header_fill, thin_border)

    for col_idx in range(1, len(consolidated_headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 18
    ws.freeze_panes = 'A3'

    ablation_excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(ablation_excel_path))
    print(f'Updated: {ablation_excel_path}')
    return True


def run_parallel_timeout_shards(
    args,
    sizes,
    algs,
    log_dir: Path | None,
):
    """
    Parent orchestration mode (one pass per puzzle size):
    - Each selected algorithm starts with ``workers_per_alg`` workers.
    - Workers use dynamic per-instance claims (work-stealing safe).
    - If one algorithm finishes early, its worker slots are transferred to the
      other algorithm to help finish remaining instances.
    - Single-algorithm mode: ``workers_per_alg`` workers, ``num_workers`` = same.
    """
    workers_per_alg = max(1, int(args.workers_per_alg))
    total_workers = workers_per_alg * max(1, len(algs))
    script_path = Path(__file__).resolve()
    best_config = str(Path(args.best_config))
    outdir = str(Path(args.outdir))
    binary = str(args.binary)

    print(
        f'[ORCHESTRATOR] Parallel timeout mode: {workers_per_alg} worker(s)/algorithm; '
        f'dynamic claims enabled (safe work-stealing); transfer enabled.'
    )

    def _spawn_alltimeouts_worker(
        *,
        size_short: str,
        size_name: str,
        alg_id: int,
        alg_name: str,
        worker_id: int,
        num_workers: int,
        stage: str,
    ) -> dict:
        session_path: Path | None = None
        if log_dir is not None:
            log_dir.mkdir(parents=True, exist_ok=True)
            sess_dir = log_dir / 'sessions'
            sess_dir.mkdir(exist_ok=True)
            safe_stage = stage.replace(' ', '_')
            session_path = (
                sess_dir / f'{size_short}_alg{alg_id}_w{worker_id}_{safe_stage}.log'
            )

        cmd = [
            sys.executable,
            str(script_path),
            '--binary', binary,
            '--best-config', best_config,
            '--outdir', outdir,
            '--reps', str(int(args.reps)),
            '--size', size_short,
            '--alg', str(int(alg_id)),
            '--worker-id', str(int(worker_id)),
            '--num-workers', str(int(num_workers)),
            '--run-all-timeouts',
            '--dynamic-claims',
            '--no-consolidate',
        ]
        if args.timeout is not None:
            cmd.extend(['--timeout', str(int(args.timeout))])
        if log_dir is not None and session_path is not None:
            cmd.extend(['--worker-session-log', str(session_path)])
        if args.no_log_files:
            cmd.append('--no-log-files')
        if args.quiet:
            cmd.append('--quiet')
        elif args.verbose:
            cmd.append('--verbose')
        else:
            cmd.append('--quiet')

        print(
            f'[ORCHESTRATOR] Spawn {alg_name} (alg={alg_id}) '
            f'worker_id={worker_id} num_workers={num_workers} [{stage}] — {size_name}'
        )
        return {
            'alg': alg_id,
            'alg_name': alg_name,
            'worker_id': worker_id,
            'num_workers': num_workers,
            'stage': stage,
            'size_name': size_name,
            'proc': subprocess.Popen(cmd, cwd=str(REPO_ROOT)),
            'session_path': session_path,
        }

    def _finalize_worker(task: dict, failed: list[tuple[str, int, int]]) -> None:
        rc = task['proc'].returncode
        if rc is None:
            rc = task['proc'].wait()
        alg = task['alg']
        alg_name = task['alg_name']
        worker_id = task['worker_id']
        stage = task['stage']
        size_name = task['size_name']
        session_path = task['session_path']
        if log_dir is not None and session_path is not None:
            _append_session_to_alg_log(
                log_dir,
                alg,
                session_path,
                f'Merged — {alg_name} (alg={alg}) worker_id={worker_id} '
                f'num_workers={task["num_workers"]} [{stage}] | {size_name} '
                f'all timeouts | exit_code={rc}',
            )
        if rc != 0:
            failed.append((alg_name, worker_id, rc))

    for size_name, _size_cfg in sizes.items():
        size_short = SIZE_SHORT[size_name]
        timeouts = TIMEOUTS_PER_SIZE.get(size_name, [])
        if args.timeout is not None:
            if args.timeout not in timeouts:
                raise SystemExit(
                    f'--timeout {args.timeout} not in grid for {size_name}: {timeouts}')

        print(
            f'\n[ORCHESTRATOR] Size phase — puzzle={size_name} '
            f'timeouts={timeouts if args.timeout is None else [args.timeout]}'
        )

        failed: list[tuple[str, int, int]] = []

        if len(algs) == 1:
            alg_id, alg_name = algs[0]
            tasks = [
                _spawn_alltimeouts_worker(
                    size_short=size_short,
                    size_name=size_name,
                    alg_id=alg_id,
                    alg_name=alg_name,
                    worker_id=w,
                    num_workers=workers_per_alg,
                    stage='single_alg',
                )
                for w in range(workers_per_alg)
            ]
            for task in tasks:
                task['proc'].wait()
                _finalize_worker(task, failed)
            if failed:
                print('[ORCHESTRATOR] ERROR: some workers failed:', file=sys.stderr)
                for alg_name_f, worker_id_f, rc in failed:
                    print(f'  {alg_name_f} worker_id={worker_id_f}: exit {rc}', file=sys.stderr)
                raise SystemExit(1)
            print(f'[ORCHESTRATOR] Size phase complete — {size_name}')
            continue

        if len(algs) != 2:
            raise SystemExit('Parallel mode with transfer expects exactly 2 algorithms.')

        alg0, alg0_name = algs[0]
        alg2, alg2_name = algs[1]
        if alg0 != 0 or alg2 != 2:
            raise SystemExit('Parallel transfer layout expects ALGORITHMS = (0, ACO), (2, CP-DCM-ACO).')

        running: list[dict] = []
        transfer_spawned_for_receiver: set[int] = set()
        for w in range(workers_per_alg):
            running.append(_spawn_alltimeouts_worker(
                size_short=size_short,
                size_name=size_name,
                alg_id=alg0,
                alg_name=alg0_name,
                worker_id=w,
                num_workers=workers_per_alg,
                stage='initial_aco',
            ))
        for w in range(workers_per_alg):
            running.append(_spawn_alltimeouts_worker(
                size_short=size_short,
                size_name=size_name,
                alg_id=alg2,
                alg_name=alg2_name,
                worker_id=w,
                num_workers=workers_per_alg,
                stage='initial_cp_dcm',
            ))

        while True:
            for task in list(running):
                if task['proc'].poll() is None:
                    continue
                running.remove(task)
                _finalize_worker(task, failed)

            if failed:
                break

            running_by_alg = {int(t['alg']) for t in running}
            transfer_specs = [
                (alg0, alg0_name, alg2, alg2_name),
                (alg2, alg2_name, alg0, alg0_name),
            ]
            for donor_alg, donor_name, receiver_alg, receiver_name in transfer_specs:
                donor_done = donor_alg not in running_by_alg
                receiver_still_running = receiver_alg in running_by_alg
                if (not donor_done) or (not receiver_still_running):
                    continue
                if receiver_alg in transfer_spawned_for_receiver:
                    continue

                transfer_spawned_for_receiver.add(receiver_alg)
                print(
                    f'[ORCHESTRATOR] All {donor_name} (alg={donor_alg}) workers finished; '
                    f'transferring worker IDs {workers_per_alg}..{total_workers - 1} '
                    f'to {receiver_name} (alg={receiver_alg}).'
                )
                for w in range(workers_per_alg, total_workers):
                    running.append(_spawn_alltimeouts_worker(
                        size_short=size_short,
                        size_name=size_name,
                        alg_id=receiver_alg,
                        alg_name=receiver_name,
                        worker_id=w,
                        num_workers=total_workers,
                        stage=f'transfer_alg{receiver_alg}_after_alg{donor_alg}',
                    ))

            if not running:
                break
            time.sleep(0.2)

        if failed:
            print('[ORCHESTRATOR] ERROR: some workers failed:', file=sys.stderr)
            for alg_name_f, worker_id_f, rc in failed:
                print(f'  {alg_name_f} worker_id={worker_id_f}: exit {rc}', file=sys.stderr)
            raise SystemExit(1)

        print(f'[ORCHESTRATOR] Size phase complete — {size_name}')


def main():
    ap = argparse.ArgumentParser(
        description='ACO vs CP-DCM-ACO across timeouts; best_config.json applies to CP-DCM-ACO only.')
    ap.add_argument('--binary', default=default_binary(), help='Solver binary path')
    ap.add_argument('--best-config', type=str, default=str(DEFAULT_BEST_CONFIG),
                    help='JSON for CP-DCM-ACO only (alg 2); ACO uses binary defaults')
    ap.add_argument('--outdir', type=str, default=str(DEFAULT_OUTDIR),
                    help='Output directory for CSVs')
    ap.add_argument('--reps', type=int, default=100, help='Repetitions per instance')
    ap.add_argument('--size', choices=['9', '16', '25'], default=None,
                    help='Single puzzle size (default: all)')
    ap.add_argument('--timeout', type=int, default=None,
                    help='Single timeout value (must belong to that size grid)')
    ap.add_argument('--alg', type=int, default=None, choices=[0, 2],
                    help='Run only this algorithm ID (default: both)')
    ap.add_argument('--workers-per-alg', type=int, default=1,
                    help='When >1, parallel parent runs all timeouts per size; '
                         'N workers per algorithm with safe work-stealing + transfer')
    ap.add_argument('--consolidate', action='store_true',
                    help='Only build timeout_comparison.xlsx from existing CSVs')
    ap.add_argument('--no-consolidate', action='store_true',
                    help='Skip Excel after runs')
    ap.add_argument('--excel-path', type=str, default=str(DEFAULT_EXCEL),
                    help='Excel output path')
    ap.add_argument('--ablation-excel-path', type=str, default=str(DEFAULT_ABLATION_EXCEL),
                    help='Ablation workbook to update with timeout tables')
    ap.add_argument('--worker-id', type=int, default=0)
    ap.add_argument('--num-workers', type=int, default=1)
    ap.add_argument(
        '--dynamic-claims',
        action='store_true',
        help='Use atomic per-instance claims instead of static modulo sharding',
    )
    ap.add_argument('--verbose', action='store_true', default=True)
    ap.add_argument('--quiet', action='store_true')
    ap.add_argument(
        '--log-dir',
        type=str,
        default=str(DEFAULT_LOG_DIR),
        help='Directory for timeout_alg0.log, timeout_alg2.log, timeout_orchestrator.log',
    )
    ap.add_argument(
        '--no-log-files',
        action='store_true',
        help='Do not write per-algorithm / orchestrator log files (stdout only)',
    )
    ap.add_argument(
        '--worker-session-log',
        default=None,
        help=argparse.SUPPRESS,
    )
    ap.add_argument(
        '--run-all-timeouts',
        action='store_true',
        help=argparse.SUPPRESS,
    )
    args = ap.parse_args()

    outdir = Path(args.outdir)
    excel_path = Path(args.excel_path)
    ablation_excel_path = Path(args.ablation_excel_path)
    verbose = args.verbose and not args.quiet

    _orig_stdout = sys.stdout
    _orch_log_f = None
    _alg_log_f = None
    _worker_sess_f = None

    log_dir: Path | None = None if args.no_log_files else Path(args.log_dir)

    def vlog(*a, **k):
        if verbose:
            print(*a, **k, flush=True)

    def _close_alg_log() -> None:
        nonlocal _alg_log_f
        if _alg_log_f is not None:
            try:
                _alg_log_f.close()
            except OSError:
                pass
            _alg_log_f = None

    def _attach_alg_log(size_name: str, alg: int, alg_name: str) -> None:
        nonlocal _alg_log_f
        if log_dir is None:
            sys.stdout = _orig_stdout
            return
        log_dir.mkdir(parents=True, exist_ok=True)
        _close_alg_log()
        p = log_dir / f'timeout_alg{alg}.log'
        _alg_log_f = open(p, 'a', encoding='utf-8', buffering=1, newline='\n')
        _alg_log_f.write('\n' + '=' * 72 + '\n')
        _alg_log_f.write(
            f'[{_log_timestamp()}] PHASE — puzzle={size_name} | {alg_name} (alg={alg}) | '
            f'serial/shard worker_id={args.worker_id} num_workers={args.num_workers}\n'
        )
        _alg_log_f.write('=' * 72 + '\n')
        _alg_log_f.flush()
        sys.stdout = _TeeStdout(_orig_stdout, _alg_log_f)

    parallel_parent = args.workers_per_alg > 1 and args.num_workers == 1
    worker_session = args.worker_session_log

    try:
        if worker_session:
            sp = Path(worker_session)
            sp.parent.mkdir(parents=True, exist_ok=True)
            _worker_sess_f = open(sp, 'w', encoding='utf-8', buffering=1, newline='\n')
            _worker_sess_f.write(
                f'[{_log_timestamp()}] Worker process — alg={args.alg} '
                f'worker_id={args.worker_id} num_workers={args.num_workers}\n'
            )
            _worker_sess_f.flush()
            sys.stdout = _TeeStdout(_orig_stdout, _worker_sess_f)
        elif log_dir is not None and (parallel_parent or args.consolidate):
            log_dir.mkdir(parents=True, exist_ok=True)
            _orch_log_f = open(
                log_dir / 'timeout_orchestrator.log',
                'a',
                encoding='utf-8',
                buffering=1,
                newline='\n',
            )
            role = 'consolidate-only' if args.consolidate else 'parallel parent'
            _orch_log_f.write('\n' + '=' * 72 + '\n')
            _orch_log_f.write(f'[{_log_timestamp()}] ORCHESTRATOR — {role}\n')
            _orch_log_f.write('=' * 72 + '\n')
            _orch_log_f.flush()
            sys.stdout = _TeeStdout(_orig_stdout, _orch_log_f)

        if args.consolidate:
            print(f'[{_log_timestamp()}] [ORCHESTRATOR] Building Excel from existing CSVs…')
            consolidate_timeout_excel(outdir, excel_path)
            merge_timeout_into_ablation_workbook(outdir, ablation_excel_path)
            return

        binary = args.binary
        if not Path(binary).exists():
            print(f'ERROR: binary not found: {binary}')
            sys.exit(1)

        size_map = {'9': '9x9', '16': '16x16', '25': '25x25'}
        if args.size:
            sizes = OrderedDict([
                (size_map[args.size], SIZE_CONFIGS[size_map[args.size]]),
            ])
        else:
            sizes = SIZE_CONFIGS

        algs = ALGORITHMS if args.alg is None else tuple(
            (a, n) for a, n in ALGORITHMS if a == args.alg)

        best_path = Path(args.best_config)
        per_size_cfg = None
        if any(a == 2 for a, _ in algs):
            per_size_cfg = load_best_config(best_path)

        # Parallel child: run every timeout for this size in one process (orchestrator spawns us).
        if getattr(args, 'run_all_timeouts', False):
            if args.size is None:
                raise SystemExit('ERROR: internal --run-all-timeouts requires --size')
            if args.alg is None:
                raise SystemExit('ERROR: internal --run-all-timeouts requires --alg')
            size_name = size_map[args.size]
            size_cfg = SIZE_CONFIGS[size_name]
            timeouts_list = list(TIMEOUTS_PER_SIZE.get(size_name, []))
            if args.timeout is not None:
                if args.timeout not in timeouts_list:
                    raise SystemExit(
                        f'--timeout {args.timeout} not in grid for {size_name}: {timeouts_list}')
                timeouts_list = [args.timeout]
            alg_map = {a: n for a, n in ALGORITHMS}
            if args.alg not in alg_map:
                raise SystemExit(f'ERROR: unknown --alg {args.alg}')
            alg_name = alg_map[args.alg]
            for idx_t, t in enumerate(timeouts_list):
                if args.alg == 0:
                    extra_args: list = []
                    vlog(
                        f'[ACO alg=0] {size_name} timeout={t}s: binary defaults '
                        f'(worker {args.worker_id}/{args.num_workers})'
                    )
                else:
                    assert per_size_cfg is not None
                    cfg = per_size_cfg[size_name]
                    extra_args, _ = build_solver_args_from_full_config(cfg)
                    vlog(
                        f'[CP-DCM-ACO alg=2] {size_name} timeout={t}s '
                        f'(worker {args.worker_id}/{args.num_workers}) config: {cfg}'
                    )
                vlog(
                    f'\n[RUN] {alg_name} (alg={args.alg}) timeout={t}s puzzle={size_name} — '
                    f'starting instances…'
                )
                run_timeout_job(
                    binary, args.alg, alg_name, t, size_name, size_cfg,
                    extra_args, args.reps, outdir, vlog,
                    worker_id=args.worker_id,
                    num_workers=args.num_workers,
                    dynamic_claims=args.dynamic_claims,
                )
                # Strict timeout ordering across workers:
                # do not start next timeout until this timeout matrix is globally complete.
                if idx_t + 1 < len(timeouts_list):
                    wait_for_timeout_matrix_complete(
                        outdir=outdir,
                        alg=args.alg,
                        timeout_sec=t,
                        size_name=size_name,
                        size_cfg=size_cfg,
                        alg_name=alg_name,
                        vlog=vlog,
                    )
            return

        if parallel_parent:
            run_parallel_timeout_shards(args, sizes, algs, log_dir)
            if not args.no_consolidate:
                vlog(
                    f'\n[{_log_timestamp()}] [ORCHESTRATOR] Consolidating timeout '
                    f'comparison Excel…'
                )
                consolidate_timeout_excel(outdir, excel_path)
                merge_timeout_into_ablation_workbook(outdir, ablation_excel_path)
            return

        for size_name, size_cfg in sizes.items():
            timeouts = TIMEOUTS_PER_SIZE.get(size_name, [])
            if args.timeout is not None:
                if args.timeout not in timeouts:
                    raise SystemExit(
                        f'--timeout {args.timeout} not in grid for {size_name}: {timeouts}')
                timeouts = [args.timeout]

            for alg, alg_name in algs:
                _attach_alg_log(size_name, alg, alg_name)
                if alg == 0:
                    extra_args: list = []
                    vlog(
                        f'[ACO alg=0] {size_name}: binary defaults '
                        f'(no best_config.json CLI overrides)'
                    )
                else:
                    assert per_size_cfg is not None
                    cfg = per_size_cfg[size_name]
                    extra_args, _ = build_solver_args_from_full_config(cfg)
                    vlog(f'[CP-DCM-ACO alg=2] {size_name} config: {cfg}')
                for t in timeouts:
                    vlog(
                        f'\n[RUN] {alg_name} (alg={alg}) timeout={t}s puzzle={size_name} — '
                        f'starting instances…'
                    )
                    run_timeout_job(
                        binary, alg, alg_name, t, size_name, size_cfg,
                        extra_args, args.reps, outdir, vlog,
                        worker_id=args.worker_id,
                        num_workers=args.num_workers,
                        dynamic_claims=args.dynamic_claims,
                    )

        if not args.no_consolidate:
            _close_alg_log()
            sys.stdout = _orig_stdout
            if log_dir is not None and not worker_session:
                log_dir.mkdir(parents=True, exist_ok=True)
                _orch_log_f = open(
                    log_dir / 'timeout_orchestrator.log',
                    'a',
                    encoding='utf-8',
                    buffering=1,
                    newline='\n',
                )
                sys.stdout = _TeeStdout(_orig_stdout, _orch_log_f)
            vlog(
                f'\n[{_log_timestamp()}] [ORCHESTRATOR] Consolidating timeout '
                f'comparison Excel…'
            )
            consolidate_timeout_excel(outdir, excel_path)
            merge_timeout_into_ablation_workbook(outdir, ablation_excel_path)

    finally:
        sys.stdout = _orig_stdout
        for fh in (_worker_sess_f, _alg_log_f, _orch_log_f):
            if fh is not None:
                try:
                    fh.close()
                except OSError:
                    pass


if __name__ == '__main__':
    main()
