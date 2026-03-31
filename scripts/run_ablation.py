#!/usr/bin/env python3
"""
Ablation study for CP-DCM-ACO (Algorithm 2).

Tests one parameter at a time while keeping all others at default values.
Results are consolidated into a single Excel file with one sheet per parameter.

Each puzzle is run 100 times (configurable via --reps) for each parameter value.
The script supports resume: if interrupted, re-running picks up where it left off.

Usage:
  python scripts/run_ablation.py                                # Run all parameters, all sizes
  python scripts/run_ablation.py --param nAnts                  # Run only the nAnts ablation
  python scripts/run_ablation.py --param nAnts --size 9         # Run nAnts on 9x9 only
  python scripts/run_ablation.py --consolidate                  # Only consolidate existing CSVs to Excel
  python scripts/run_ablation.py --reps 10 --param q0           # Quick test with 10 reps

Requirements:
  pip install openpyxl
"""

import argparse
import csv
import math
import os
import sys
import time
from collections import OrderedDict
from pathlib import Path

try:
    import msvcrt
    HAS_MSVCRT = True
except ImportError:
    HAS_MSVCRT = False
    try:
        import fcntl
        HAS_FCNTL = True
    except ImportError:
        HAS_FCNTL = False

from bench_utils import default_binary, run_solver, safe_mean, safe_std

# ============================================================
# Configuration
# ============================================================

ALG = 2
ALG_NAME = 'CP-DCM-ACO'

DEFAULTS = {
    'nAnts': 3,
    'numACS': 2,
    'q0': 0.9,
    'xi': 0.1,
    'rho': 0.9,
    'evap': 0.005,
    'convThresh': 0.8,
    'entropyPct': 92.5,
}

SIZE_CONFIGS = OrderedDict([
    ('9x9',  {'dir': 'instances/9x9',  'timeout': 5}),
    ('16x16', {'dir': 'instances/16x16', 'timeout': 20}),
    ('25x25', {'dir': 'instances/25x25', 'timeout': 120}),
])

PARAM_TESTS = OrderedDict([
    ('nAnts', {
        'label': 'Number of Ants (Per Colony)',
        'values': [5, 7, 9, 11],
    }),
    ('numACS', {
        'label': 'Number of ACS Colonies',
        'values': [3, 4, 5, 6],
    }),
    ('q0', {
        'label': 'q0 (Greedy vs Roulette)',
        'values': [0.3, 0.5, 0.7, 0.99],
    }),
    ('xi', {
        'label': 'xi (Local Pheromone Update)',
        'values': [0.01, 0.3, 0.5, 0.7],
    }),
    ('rho', {
        'label': 'rho (Pheromone Persistence)',
        'values': [0.3, 0.5, 0.7, 0.99],
    }),
    ('evap', {
        'label': 'BVE (Evaporation Rate of BestPher)',
        'values': [0.0025, 0.0075, 0.01, 0.0125],
    }),
    ('convThresh', {
        'label': 'Convergence Threshold',
        'values': [0.2, 0.4, 0.6, 1.0],
    }),
    ('entropyPct', {
        'label': 'Entropy Threshold (% of Max Entropy)',
        'values': [78.625, 83.25, 87.875, 97.125],
    }),
    ('timeout', {
        'label': 'Timeout (seconds)',
        'values_per_size': {
            '9x9':  [1, 3, 7, 9],
            '16x16': [10, 15, 25, 30],
            '25x25': [60, 90, 150, 180],
        },
    }),
])

ABLATION_DIR = Path('results') / 'ablation'

SUMMARY_HEADERS = [
    'param_value', 'puzzle_size', 'instance',
    'alg', 'alg_name',
    'success_%', 'time_mean', 'time_std', 'cycles_mean', 'cycles_std',
]

PROGRESS_HEADERS = [
    'instance', 'rep', 'success', 'time', 'cycles',
]


# ============================================================
# Helpers
# ============================================================

def max_entropy(n_ants):
    """Theoretical maximum Shannon entropy: log2(n_ants)."""
    return math.log2(n_ants)


def compute_entropy_threshold(n_ants, pct):
    """Absolute entropy threshold from percentage of max entropy."""
    return max_entropy(n_ants) * (pct / 100.0)


def lock_file(fh):
    try:
        if HAS_MSVCRT:
            sz = os.path.getsize(fh.name) if hasattr(fh, 'name') else 1
            msvcrt.locking(fh.fileno(), msvcrt.LK_LOCK, max(1, sz))
        elif HAS_FCNTL:
            fcntl.flock(fh.fileno(), fcntl.LOCK_EX)
    except Exception:
        pass


def unlock_file(fh):
    try:
        if HAS_MSVCRT:
            sz = os.path.getsize(fh.name) if hasattr(fh, 'name') else 1
            msvcrt.locking(fh.fileno(), msvcrt.LK_UNLCK, max(1, sz))
        elif HAS_FCNTL:
            fcntl.flock(fh.fileno(), fcntl.LOCK_UN)
    except Exception:
        pass


def scan_instances(instances_dir):
    return sorted(Path(instances_dir).glob('*.txt'))


def ensure_csv_header(path, headers):
    if path.exists():
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, 'w', newline='') as f:
        csv.writer(f).writerow(headers)


def append_csv_row(path, row):
    retries = 10
    for attempt in range(retries):
        try:
            with open(path, 'a', newline='') as f:
                lock_file(f)
                csv.writer(f).writerow(row)
                unlock_file(f)
            return True
        except (IOError, OSError):
            if attempt < retries - 1:
                time.sleep(0.1 * (attempt + 1))
    return False


def read_progress(progress_file):
    """Read per-rep progress → dict[instance_name → dict[rep → (success, time, cycles)]]"""
    prog = {}
    if not progress_file.exists():
        return prog
    try:
        with open(progress_file, 'r', newline='') as f:
            lock_file(f)
            for row in csv.DictReader(f):
                inst = row.get('instance')
                rep_s = row.get('rep')
                if not inst or not rep_s:
                    continue
                try:
                    rep = int(rep_s)
                except ValueError:
                    continue
                success = row.get('success', '').strip() in ('1', 'true', 'True')
                t = math.nan
                cyc = math.nan
                try:
                    if row.get('time', '').strip():
                        t = float(row['time'])
                except Exception:
                    pass
                try:
                    if row.get('cycles', '').strip():
                        cyc = float(row['cycles'])
                except Exception:
                    pass
                prog.setdefault(inst, {})[rep] = (success, t, cyc)
            unlock_file(f)
    except Exception:
        return {}
    return prog


def read_completed_from_summary(summary_file):
    completed = set()
    if not summary_file.exists():
        return completed
    try:
        with open(summary_file, 'r', newline='') as f:
            lock_file(f)
            for row in csv.DictReader(f):
                inst = row.get('instance')
                if inst:
                    completed.add(inst)
            unlock_file(f)
    except Exception:
        return set()
    return completed


def size_name_from_summary_filename(filename: str):
    """Parse e.g. ``6_25x25_summary.csv`` or ``0.99_25x25_summary.csv`` → ``25x25``."""
    if not filename.endswith('_summary.csv'):
        return None
    base = filename[: -len('_summary.csv')]
    for sz in ('25x25', '16x16', '9x9'):
        if base.endswith('_' + sz):
            return sz
    return None


def sort_summary_csv_if_complete(summary_file: Path, canonical_instance_names: list) -> bool:
    """If the summary has one row per canonical instance, rewrite rows in instance-folder order.

    Parallel workers append rows in completion order; this restores the same ordering as
    ``sorted(instances/*.txt)`` for thesis tables and comparison with ``numACS`` single-worker runs.
    """
    if not summary_file.exists() or not canonical_instance_names:
        return False
    canonical_set = set(canonical_instance_names)
    try:
        with open(summary_file, 'r+', newline='') as f:
            lock_file(f)
            try:
                reader = csv.DictReader(f)
                fieldnames = reader.fieldnames
                if not fieldnames:
                    return False
                rows = list(reader)
                if len(rows) != len(canonical_instance_names):
                    return False
                by_inst = {}
                for row in rows:
                    inst = (row.get('instance') or '').strip()
                    if inst:
                        by_inst[inst] = row
                if set(by_inst.keys()) != canonical_set:
                    return False
                f.seek(0)
                f.truncate()
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                for name in canonical_instance_names:
                    writer.writerow(by_inst[name])
                f.flush()
                try:
                    os.fsync(f.fileno())
                except OSError:
                    pass
            finally:
                unlock_file(f)
        return True
    except Exception:
        return False


def format_param_value(param_name, value):
    """Produce a clean string for file names and display."""
    if isinstance(value, float):
        s = f'{value:g}'
    else:
        s = str(value)
    return s


def build_solver_args(param_name, param_value):
    """Build extra CLI args for the solver and return (args_list, entropy_threshold_used)."""
    n_ants = DEFAULTS['nAnts']
    entropy_pct = DEFAULTS['entropyPct']

    if param_name == 'nAnts':
        n_ants = int(param_value)
    elif param_name == 'entropyPct':
        entropy_pct = float(param_value)

    args = []
    args += ['--nAnts', str(n_ants)]

    num_acs = int(param_value) if param_name == 'numACS' else DEFAULTS['numACS']
    args += ['--numACS', str(num_acs), '--numColonies', str(num_acs + 1)]

    q0 = float(param_value) if param_name == 'q0' else DEFAULTS['q0']
    args += ['--q0', str(q0)]

    xi = float(param_value) if param_name == 'xi' else DEFAULTS['xi']
    args += ['--xi', str(xi)]

    rho = float(param_value) if param_name == 'rho' else DEFAULTS['rho']
    args += ['--rho', str(rho)]

    evap = float(param_value) if param_name == 'evap' else DEFAULTS['evap']
    args += ['--evap', str(evap)]

    conv = float(param_value) if param_name == 'convThresh' else DEFAULTS['convThresh']
    args += ['--convThresh', str(conv)]

    ent_thresh = compute_entropy_threshold(n_ants, entropy_pct)
    args += ['--entropyThreshold', str(round(ent_thresh, 6))]

    return args, round(ent_thresh, 6)


def delete_ablation_progress_if_summary_done(
    progress_file: Path,
    summary_file: Path,
    all_instance_names: set,
    vlog,
    tag: str,
):
    """Delete progress CSV when the summary lists every expected instance."""
    if not summary_file.exists() or not all_instance_names:
        return
    completed = read_completed_from_summary(summary_file)
    if not all_instance_names <= completed:
        return
    if not progress_file.exists():
        return
    try:
        progress_file.unlink()
        vlog(f'  [{tag}] Summary complete; removed {progress_file.name}')
    except OSError as e:
        vlog(f'  [{tag}] WARNING: could not remove progress file: {e}')


# ============================================================
# Core test runner
# ============================================================

def run_ablation_test(binary, param_name, param_value, size_name, size_cfg,
                      reps, outdir, vlog, timeout_override=None,
                      worker_id: int = 0, num_workers: int = 1):
    """Run all instances for one (param, value, size) combo. Returns summary rows."""

    val_str = format_param_value(param_name, param_value)
    tag = f'{param_name}={val_str} [{size_name}]'

    param_dir = outdir / param_name
    param_dir.mkdir(parents=True, exist_ok=True)

    progress_file = param_dir / f'{val_str}_{size_name}_progress.csv'
    summary_file = param_dir / f'{val_str}_{size_name}_summary.csv'

    # Build solver args (all defaults except the tested param)
    if param_name == 'timeout':
        extra_args, ent_thresh = build_solver_args(None, None)
        timeout = int(param_value)
    else:
        extra_args, ent_thresh = build_solver_args(param_name, param_value)
        timeout = timeout_override if timeout_override else size_cfg['timeout']

    instances_all = scan_instances(size_cfg['dir'])
    if not instances_all:
        vlog(f'  No instances found in {size_cfg["dir"]}')
        return []

    all_instance_names = {fp.name for fp in instances_all}

    # Partition instances deterministically across workers so each worker handles
    # a disjoint subset of instances (prevents duplicate summary rows).
    if num_workers < 1:
        num_workers = 1
    worker_id = int(worker_id) % int(num_workers)
    instances = [
        fp for i, fp in enumerate(instances_all)
        if (i % num_workers) == worker_id
    ]

    if not instances:
        vlog(f'  Worker {worker_id}/{num_workers} has no assigned instances')
        return []

    completed = read_completed_from_summary(summary_file)

    # If the summary is already complete for *all* instances, we never need
    # progress anymore. Also delete any existing progress file so it does not
    # get recreated by this job.
    completed_overall = len(completed.intersection(all_instance_names))
    if completed_overall == len(instances_all):
        sort_summary_csv_if_complete(summary_file, [fp.name for fp in instances_all])
        delete_ablation_progress_if_summary_done(
            progress_file, summary_file, all_instance_names, vlog, tag)
        vlog(f'  [{tag}] Already complete ({completed_overall}/{len(instances_all)}). '
             f'Skipping.')
        return []

    progress = read_progress(progress_file)

    ensure_csv_header(summary_file, SUMMARY_HEADERS)
    # Only create progress.csv when we actually need to resume/continue work.
    ensure_csv_header(progress_file, PROGRESS_HEADERS)

    subset_names = {fp.name for fp in instances}
    completed_in_subset = completed.intersection(subset_names)

    total = len(instances)
    summary_rows = []

    if completed_in_subset:
        vlog(f'  [{tag}] Worker {worker_id}/{num_workers} Resuming: '
             f'{len(completed_in_subset)}/{total} instances already done')

    for idx, fp in enumerate(instances, 1):
        if fp.name in completed:
            continue

        rep_map = progress.get(fp.name, {})
        done_reps = set(rep_map.keys())

        successes = 0
        times = []
        cycles_solved = []
        for _r, (succ, t, cyc) in sorted(rep_map.items()):
            if succ:
                successes += 1
                times.append(t)
                if not math.isnan(cyc):
                    cycles_solved.append(cyc)

        status = f'RESUME {len(done_reps)}/{reps}' if done_reps else ''
        vlog(f'  [{tag}] ({idx}/{total}) {fp.name} {status}')

        for rep in range(1, reps + 1):
            if rep in done_reps:
                continue

            success, t, cyc, out = run_solver(binary, fp, ALG, timeout, extra_args=extra_args)

            status_str = 'OK' if success else 'FAIL'
            t_str = f'{t:.4f}s' if not math.isnan(t) else 'N/A'
            vlog(f'    Rep {rep}/{reps}: {status_str} (time={t_str})')

            prog_row = [fp.name, rep, 1 if success else 0,
                        '' if math.isnan(t) else t,
                        '' if math.isnan(cyc) else cyc]
            append_csv_row(progress_file, prog_row)

            rep_map[rep] = (success, t, cyc)
            done_reps.add(rep)
            if success:
                successes += 1
                times.append(t)
                if not math.isnan(cyc):
                    cycles_solved.append(cyc)

        if len(done_reps) < reps:
            vlog(f'    => partial ({len(done_reps)}/{reps})')
            progress[fp.name] = rep_map
            continue

        succ_pct = (successes / float(reps)) * 100.0
        tm = safe_mean(times)
        ts = safe_std(times)
        cm = safe_mean(cycles_solved)
        cs = safe_std(cycles_solved)

        row = [
            param_value, size_name, fp.name,
            ALG, ALG_NAME,
            round(succ_pct, 2),
            round(tm, 6) if not math.isnan(tm) else '',
            round(ts, 6) if not math.isnan(ts) else '',
            round(cm, 3) if not math.isnan(cm) else '',
            round(cs, 3) if not math.isnan(cs) else '',
        ]

        if append_csv_row(summary_file, row):
            completed.add(fp.name)
            summary_rows.append(row)
            completed_now = read_completed_from_summary(summary_file)
            if all_instance_names <= completed_now:
                sort_summary_csv_if_complete(summary_file, [p.name for p in instances_all])
                delete_ablation_progress_if_summary_done(
                    progress_file, summary_file, all_instance_names, vlog, tag)
            vlog(f'    => success%={round(succ_pct,2)} '
                 f'time_mean={round(tm,6) if not math.isnan(tm) else "N/A"} '
                 f'cycles_mean={round(cm,3) if not math.isnan(cm) else "N/A"}')
        else:
            vlog(f'    ERROR: Could not write summary for {fp.name}')

        progress[fp.name] = rep_map

    # Other worker may have finished last; delete progress if summary is now full.
    delete_ablation_progress_if_summary_done(
        progress_file, summary_file, all_instance_names, vlog, tag)

    return summary_rows


# ============================================================
# Excel consolidation
# ============================================================

def consolidate_to_excel(outdir, excel_path):
    """Read all summary CSVs and write a single Excel workbook."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        print('ERROR: openpyxl is required for Excel output.  pip install openpyxl')
        return False

    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    default_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    for param_name, pcfg in PARAM_TESTS.items():
        param_dir = outdir / param_name
        if not param_dir.exists():
            continue

        all_rows = []
        for csv_file in sorted(param_dir.glob('*_summary.csv')):
            sz = size_name_from_summary_filename(csv_file.name)
            if sz:
                canon = [fp.name for fp in scan_instances(Path('instances') / sz)]
                if canon:
                    sort_summary_csv_if_complete(csv_file, canon)
                    canon_set = set(canon)
                    if canon_set <= read_completed_from_summary(csv_file):
                        prog_path = csv_file.parent / csv_file.name.replace(
                            '_summary.csv', '_progress.csv')
                        if prog_path.exists():
                            try:
                                prog_path.unlink()
                            except OSError:
                                pass
            try:
                with open(csv_file, 'r', newline='') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        all_rows.append(row)
            except Exception:
                continue

        if not all_rows:
            continue

        label = pcfg['label']
        sheet_name = label[:31]  # Excel sheet name limit
        ws = wb.create_sheet(title=sheet_name)

        headers = [
            pcfg['label'], 'Puzzle Size', 'Instance',
            'Algorithm', 'Algorithm Name',
            'Success %', 'Time Mean (s)', 'Time Std (s)',
            'Cycles Mean', 'Cycles Std',
        ]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border

        # Determine which param values are default
        if param_name == 'timeout':
            default_vals = set()
            vps = pcfg.get('values_per_size', {})
            for sz, cfg in SIZE_CONFIGS.items():
                default_vals.add(str(cfg['timeout']))
        elif param_name in DEFAULTS:
            default_vals = {str(DEFAULTS[param_name])}
        else:
            default_vals = set()

        for r_idx, row in enumerate(all_rows, 2):
            vals = [
                row.get('param_value', ''),
                row.get('puzzle_size', ''),
                row.get('instance', ''),
                row.get('alg', ''),
                row.get('alg_name', ''),
                row.get('success_%', ''),
                row.get('time_mean', ''),
                row.get('time_std', ''),
                row.get('cycles_mean', ''),
                row.get('cycles_std', ''),
            ]
            for col_idx, v in enumerate(vals, 1):
                try:
                    v_num = float(v) if v != '' else v
                    if v_num == int(v_num) and col_idx not in (7, 8, 9, 10):
                        v_num = int(v_num)
                    cell = ws.cell(row=r_idx, column=col_idx, value=v_num)
                except (ValueError, TypeError, OverflowError):
                    cell = ws.cell(row=r_idx, column=col_idx, value=v)
                cell.border = thin_border

                is_default = str(row.get('param_value', '')) in default_vals
                if is_default:
                    cell.fill = default_fill

        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 18
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = 'A2'

    if not wb.sheetnames:
        print('No ablation data found to consolidate.')
        return False

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(excel_path))
    print(f'Excel file saved: {excel_path}')
    return True


# ============================================================
# Main
# ============================================================

def main():
    ap = argparse.ArgumentParser(
        description='Ablation study for CP-DCM-ACO. Tests one parameter at a time, '
                    'keeping all others at default. Results are consolidated into Excel.')
    ap.add_argument('--binary', default=default_binary(),
                    help='Path to solver binary (default: auto-detect)')
    ap.add_argument('--param', type=str, default=None,
                    choices=list(PARAM_TESTS.keys()),
                    help='Run only this parameter (default: all)')
    ap.add_argument('--param-value', type=str, default=None,
                    help='Run only this parameter value for --param (default: all values)')
    ap.add_argument('--size', type=str, default=None,
                    choices=['9', '16', '25'],
                    help='Run only this puzzle size: 9, 16, or 25 (default: all)')
    ap.add_argument('--reps', type=int, default=100,
                    help='Repetitions per instance (default: 100)')
    ap.add_argument('--outdir', default=str(ABLATION_DIR),
                    help=f'Output directory (default: {ABLATION_DIR})')
    ap.add_argument('--consolidate', action='store_true',
                    help='Only consolidate existing CSVs to Excel (no tests)')
    ap.add_argument('--verbose', action='store_true', default=True,
                    help='Print progress (default: True)')
    ap.add_argument('--quiet', action='store_true',
                    help='Suppress progress output')
    ap.add_argument('--no-consolidate', action='store_true',
                    help='Skip Excel consolidation at the end (useful for parallel runs)')
    ap.add_argument('--worker-id', type=int, default=0,
                    help='Worker index for partitioning instance set (0-based)')
    ap.add_argument('--num-workers', type=int, default=1,
                    help='Number of workers partitioning one (param,value,size) job')
    args = ap.parse_args()

    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)
    excel_path = outdir / 'ablation_results.xlsx'

    verbose = args.verbose and not args.quiet

    def vlog(*a, **k):
        if verbose:
            print(*a, **k, flush=True)

    if args.consolidate:
        consolidate_to_excel(outdir, excel_path)
        return

    binary = args.binary
    if not Path(binary).exists():
        print(f'ERROR: Solver binary not found: {binary}')
        print('Please build the solver first, or specify --binary.')
        sys.exit(1)

    size_map = {'9': '9x9', '16': '16x16', '25': '25x25'}
    if args.size:
        sizes_to_test = OrderedDict([(size_map[args.size], SIZE_CONFIGS[size_map[args.size]])])
    else:
        sizes_to_test = SIZE_CONFIGS

    params_to_test = OrderedDict()
    if args.param:
        params_to_test[args.param] = PARAM_TESTS[args.param]
    else:
        params_to_test = PARAM_TESTS

    filtered_single_param_value = None
    if args.param_value is not None:
        if not args.param:
            raise SystemExit('--param-value requires --param')

        # Parse and match provided value against the configured candidate list.
        candidates = PARAM_TESTS[args.param]['values']
        parsed = None
        try:
            raw_num = float(args.param_value)
            for c in candidates:
                try:
                    if abs(float(c) - raw_num) <= 1e-12:
                        parsed = c
                        break
                except Exception:
                    continue
        except Exception:
            parsed = None

        if parsed is None:
            for c in candidates:
                if str(c) == args.param_value:
                    parsed = c
                    break

        if parsed is None:
            raise SystemExit(f'Invalid --param-value={args.param_value!r} for --param={args.param!r}')
        filtered_single_param_value = parsed

    total_configs = 0
    for pname, pcfg in params_to_test.items():
        if pname == 'timeout':
            for sname in sizes_to_test:
                vals = pcfg.get('values_per_size', {}).get(sname, [])
                if filtered_single_param_value is not None and pname == args.param:
                    total_configs += 1 if any(abs(float(v) - float(filtered_single_param_value)) <= 1e-12 for v in vals) else 0
                else:
                    total_configs += len(vals)
        else:
            if filtered_single_param_value is not None and pname == args.param:
                total_configs += 1 * len(sizes_to_test)
            else:
                total_configs += len(pcfg['values']) * len(sizes_to_test)

    vlog(f'{"="*70}')
    vlog(f'Ablation Study for {ALG_NAME}')
    vlog(f'Parameters to test: {", ".join(params_to_test.keys())}')
    vlog(f'Puzzle sizes: {", ".join(sizes_to_test.keys())}')
    vlog(f'Repetitions per instance: {args.reps}')
    vlog(f'Total configurations: {total_configs}')
    vlog(f'Output directory: {outdir}')
    vlog(f'{"="*70}')

    worker_id = int(args.worker_id)
    num_workers = int(args.num_workers)

    config_idx = 0
    for size_name, size_cfg in sizes_to_test.items():
        vlog(f'\n{"="*70}')
        vlog(f'Puzzle size: {size_name}')
        vlog(f'{"="*70}')

        for param_name, pcfg in params_to_test.items():
            vlog(f'\n  Parameter: {pcfg["label"]} ({param_name})')

            if param_name == 'timeout':
                timeout_vals = pcfg.get('values_per_size', {}).get(size_name, [])
                if filtered_single_param_value is not None and param_name == args.param:
                    timeout_vals_iter = [
                        tval for tval in timeout_vals
                        if abs(float(tval) - float(filtered_single_param_value)) <= 1e-12
                    ]
                else:
                    timeout_vals_iter = timeout_vals

                for tval in timeout_vals_iter:
                    config_idx += 1
                    vlog(f'\n[Config {config_idx}/{total_configs}] '
                         f'timeout={tval}s on {size_name}')
                    run_ablation_test(
                        binary, 'timeout', tval, size_name, size_cfg,
                        args.reps, outdir, vlog,
                        worker_id=worker_id, num_workers=num_workers)
            else:
                if filtered_single_param_value is not None and param_name == args.param:
                    values_iter = [filtered_single_param_value]
                else:
                    values_iter = pcfg['values']

                for value in values_iter:
                    config_idx += 1
                    val_str = format_param_value(param_name, value)
                    vlog(f'\n[Config {config_idx}/{total_configs}] '
                         f'{param_name}={val_str} on {size_name}')
                    run_ablation_test(
                        binary, param_name, value, size_name, size_cfg,
                        args.reps, outdir, vlog,
                        worker_id=worker_id, num_workers=num_workers)

    if not args.no_consolidate:
        vlog(f'\n{"="*70}')
        vlog('All ablation tests completed. Consolidating to Excel...')
        vlog(f'{"="*70}')
        consolidate_to_excel(outdir, excel_path)

        vlog(f'\nDone! Results saved to: {excel_path}')


if __name__ == '__main__':
    main()
