#!/usr/bin/env python3
"""
Run CP comparison benchmark runs for:
- CP-ACS (alg 0)
- CP-DCM-ACO (alg 2)

Execution model (defaults):
- Runs 1..5
- Repetitions per instance: 1
- For each size: finish requested runs, then move to next size
- Size order: 9x9 -> 16x16 -> 25x25
- For each size: start alg 0 and alg 2 in parallel
- Each algorithm uses pool workers (default: 4)
- After CP comparison phase, run extra DCM-ACO 9-ants phase:
  alg 2 only, nAnts=3, numACS=2 (thus 3 colonies = 2 ACS + 1 MMAS),
  reps=1, runs 1..5, same size order.

This script orchestrates existing benchmark scripts:
- scripts/run_9x9.py
- scripts/run_16x16.py
- scripts/run_25x25.py
"""

from __future__ import annotations

import argparse
import csv
import math
import subprocess
import sys
import time
from pathlib import Path


SIZES = [
    ("9x9", "scripts/run_9x9.py"),
    ("16x16", "scripts/run_16x16.py"),
    ("25x25", "scripts/run_25x25.py"),
]
SIZE_ALIASES = {
    "9": "9x9",
    "16": "16x16",
    "25": "25x25",
    "9x9": "9x9",
    "16x16": "16x16",
    "25x25": "25x25",
}

ALGS = [
    (0, "CP-ACS"),
    (2, "CP-DCM-ACO"),
]


def _now() -> str:
    return time.strftime("%Y-%m-%d %H:%M:%S")


def _build_cmd(
    python_exe: str,
    script_rel: str,
    alg: int,
    run_idx: int,
    reps: int,
    pool_workers: int,
    outdir: str | None,
    extra_solver_args: list[str] | None,
    binary: str | None,
    verbose: bool,
) -> list[str]:
    cmd = [
        python_exe,
        script_rel,
        "--alg",
        str(int(alg)),
        "--run",
        str(int(run_idx)),
        "--reps",
        str(int(reps)),
        "--pool-workers",
        str(int(pool_workers)),
    ]
    if outdir:
        cmd.extend(["--outdir", outdir])
    if extra_solver_args:
        cmd.extend(extra_solver_args)
    if binary:
        cmd.extend(["--binary", binary])
    if verbose:
        cmd.append("--verbose")
    return cmd


def _terminate(proc: subprocess.Popen, timeout_s: float = 5.0) -> None:
    if proc.poll() is not None:
        return
    proc.terminate()
    try:
        proc.wait(timeout=timeout_s)
    except subprocess.TimeoutExpired:
        proc.kill()
        proc.wait()


def _run_size_phase(
    *,
    repo_root: Path,
    python_exe: str,
    size_name: str,
    script_rel: str,
    run_idx: int,
    reps: int,
    pool_workers: int,
    outdir: str | None,
    extra_solver_args: list[str] | None,
    binary: str | None,
    algs: list[tuple[int, str]],
    enable_transfer: bool,
    verbose: bool,
    dry_run: bool,
) -> int:
    cmds = []
    cmds_by_alg: dict[int, tuple[str, list[str]]] = {}
    for alg, alg_name in algs:
        cmd = _build_cmd(
            python_exe=python_exe,
            script_rel=script_rel,
            alg=alg,
            run_idx=run_idx,
            reps=reps,
            pool_workers=pool_workers,
            outdir=outdir,
            extra_solver_args=extra_solver_args,
            binary=binary,
            verbose=verbose,
        )
        cmds.append((alg, alg_name, cmd))
        cmds_by_alg[alg] = (alg_name, cmd)

    print(
        f"[{_now()}] [RUN {run_idx}] [SIZE {size_name}] "
        f"starting {len(cmds)} algorithm processes in parallel",
        flush=True,
    )
    for _, alg_name, cmd in cmds:
        print(f"  - {alg_name}: {' '.join(cmd)}", flush=True)

    if dry_run:
        return 0

    transfer_allowed = bool(enable_transfer and len(algs) == 2)
    transfer_done = False
    transfer_added_workers = max(1, int(pool_workers))

    procs: dict[int, tuple[str, subprocess.Popen]] = {}
    for alg, alg_name, cmd in cmds:
        procs[alg] = (
            alg_name,
            subprocess.Popen(cmd, cwd=str(repo_root)),
        )

    first_failure: tuple[int, str, int] | None = None
    try:
        while procs:
            completed = []
            for alg, (alg_name, proc) in procs.items():
                rc = proc.poll()
                if rc is None:
                    continue
                completed.append(alg)
                print(
                    f"[{_now()}] [RUN {run_idx}] [SIZE {size_name}] "
                    f"{alg_name} finished with exit_code={rc}",
                    flush=True,
                )
                if rc != 0 and first_failure is None:
                    first_failure = (alg, alg_name, rc)

            for alg in completed:
                procs.pop(alg, None)

            if transfer_allowed and not transfer_done and first_failure is None and len(procs) == 1:
                # One algorithm finished successfully; transfer its worker slots by restarting
                # the remaining algorithm with a larger pool size and resume from progress CSVs.
                remaining_alg = next(iter(procs.keys()))
                remaining_name, remaining_proc = procs[remaining_alg]
                _terminate(remaining_proc)
                procs.pop(remaining_alg, None)

                new_workers = int(pool_workers) + transfer_added_workers
                _name, base_cmd = cmds_by_alg[remaining_alg]
                new_cmd = list(base_cmd)
                try:
                    idx = new_cmd.index("--pool-workers")
                    new_cmd[idx + 1] = str(new_workers)
                except (ValueError, IndexError):
                    # Fallback: if flag is missing for any reason, append it.
                    new_cmd.extend(["--pool-workers", str(new_workers)])

                print(
                    f"[{_now()}] [RUN {run_idx}] [SIZE {size_name}] transfer: "
                    f"restarting {remaining_name} with --pool-workers {new_workers}",
                    flush=True,
                )
                procs[remaining_alg] = (
                    remaining_name,
                    subprocess.Popen(new_cmd, cwd=str(repo_root)),
                )
                transfer_done = True
                continue

            if first_failure is not None:
                # Stop peers if one process failed.
                for _, proc in procs.values():
                    _terminate(proc)
                procs.clear()
                break

            if procs:
                time.sleep(0.25)
    except KeyboardInterrupt:
        print(f"[{_now()}] Interrupted. Terminating running processes...", flush=True)
        for _, proc in procs.values():
            _terminate(proc)
        raise

    if first_failure is not None:
        _alg, alg_name, rc = first_failure
        print(
            f"[{_now()}] [RUN {run_idx}] [SIZE {size_name}] "
            f"FAILED due to {alg_name} exit_code={rc}",
            flush=True,
        )
        return int(rc)

    print(
        f"[{_now()}] [RUN {run_idx}] [SIZE {size_name}] completed",
        flush=True,
    )
    return 0


def _default_outdir(size_name: str) -> str:
    return str(Path("results") / size_name)


def _dcm9_outdir(size_name: str) -> str:
    return str(Path("results") / size_name / "dcm_9ants")


def _safe_mean(values: list[float]) -> float:
    nums = [v for v in values if v is not None and not math.isnan(v)]
    if not nums:
        return math.nan
    return sum(nums) / float(len(nums))


def _safe_std(values: list[float]) -> float:
    nums = [v for v in values if v is not None and not math.isnan(v)]
    n = len(nums)
    if n < 2:
        return math.nan
    m = _safe_mean(nums)
    var = sum((x - m) ** 2 for x in nums) / float(n - 1)
    return math.sqrt(var)


def _parse_float(value: str | None) -> float:
    if value is None:
        return math.nan
    s = str(value).strip()
    if not s:
        return math.nan
    try:
        return float(s)
    except Exception:
        return math.nan


def _run_file_for(outdir: Path, size_name: str, alg_name: str, run_idx: int) -> Path:
    if int(run_idx) == 1:
        return outdir / f"results_{size_name}_{alg_name}.csv"
    return outdir / f"results_{size_name}_{alg_name}_run{int(run_idx)}.csv"


def _load_run_metrics(
    csv_path: Path,
    expected_instance_names: set[str] | None = None,
) -> tuple[float, float, float, bool] | None:
    if not csv_path.exists():
        return None
    succ_by_instance: dict[str, float] = {}
    tmean = []
    cycmean = []
    try:
        with open(csv_path, "r", newline="") as f:
            for row in csv.DictReader(f):
                inst = str(row.get("instance") or "").strip()
                sv = _parse_float(row.get("success_%"))
                if inst:
                    succ_by_instance[inst] = sv
                tmean.append(_parse_float(row.get("time_mean")))
                cycmean.append(_parse_float(row.get("cycles_mean")))
    except Exception:
        return None
    if not succ_by_instance and not tmean and not cycmean:
        return None
    # Determine whether this run has full instance coverage.
    is_complete = True
    if expected_instance_names:
        is_complete = expected_instance_names.issubset(set(succ_by_instance.keys()))

    # Success rate for the run is based on rows recorded in that run file.
    # A failed instance row (success_% = 0) is naturally counted in this mean.
    run_success = _safe_mean(list(succ_by_instance.values()))
    return (run_success, _safe_mean(tmean), _safe_mean(cycmean), is_complete)


def _aggregate_alg_over_runs(
    *,
    outdir: Path,
    size_name: str,
    alg_name: str,
    run_start: int,
    run_end: int,
    expected_instance_names: set[str] | None = None,
) -> dict:
    run_success = []
    run_time = []
    run_cycles = []
    found_files = 0
    incomplete_runs = 0
    for run_idx in range(int(run_start), int(run_end) + 1):
        fp = _run_file_for(outdir, size_name, alg_name, run_idx)
        m = _load_run_metrics(fp, expected_instance_names=expected_instance_names)
        if m is None:
            continue
        s, t, c, is_complete = m
        found_files += 1
        if expected_instance_names and not is_complete:
            incomplete_runs += 1
        run_success.append(s)
        run_time.append(t)
        run_cycles.append(c)

    if not run_success:
        return {
            "files_found": found_files,
            "runs_used": 0,
            "runs_incomplete": incomplete_runs,
            "best_success": math.nan,
            "worst_success": math.nan,
            "avg_success": math.nan,
            "time_mean": math.nan,
            "time_std": math.nan,
            "cycles_mean": math.nan,
        }

    clean_s = [v for v in run_success if not math.isnan(v)]
    best_s = max(clean_s) if clean_s else math.nan
    worst_s = min(clean_s) if clean_s else math.nan
    avg_s = _safe_mean(run_success)
    time_m = _safe_mean(run_time)
    time_s = _safe_std(run_time)
    cyc_m = _safe_mean(run_cycles)
    return {
        "files_found": found_files,
        "runs_used": len(run_success),
        "runs_incomplete": incomplete_runs,
        "best_success": best_s,
        "worst_success": worst_s,
        "avg_success": avg_s,
        "time_mean": time_m,
        "time_std": time_s,
        "cycles_mean": cyc_m,
    }


def _expected_instances(repo_root: Path, size_name: str) -> set[str]:
    inst_dir = repo_root / "instances" / size_name
    return {p.name for p in inst_dir.glob("*.txt")}


def _fmt(value: float, ndigits: int = 4):
    if value is None or math.isnan(value):
        return ""
    return round(float(value), int(ndigits))


def _write_sheet_tables(ws, title: str, rows_by_size: list[tuple[str, list[list]]]) -> None:
    ws.cell(row=1, column=1, value=title)
    r = 3
    headers = [
        "Algorithm",
        "Best success %",
        "Worst success %",
        "Average success %",
        "Time mean (s)",
        "Time std (s)",
        "Cycles mean",
    ]
    for size_name, rows in rows_by_size:
        ws.cell(row=r, column=1, value=f"Puzzle size: {size_name}")
        r += 1
        for c, h in enumerate(headers, start=1):
            ws.cell(row=r, column=c, value=h)
        r += 1
        for row_vals in rows:
            for c, v in enumerate(row_vals, start=1):
                ws.cell(row=r, column=c, value=v)
            r += 1
        r += 1


def consolidate_excel(
    *,
    repo_root: Path,
    selected_sizes: list[tuple[str, str]],
    run_start: int,
    run_end: int,
    excel_path: Path,
) -> bool:
    try:
        from openpyxl import Workbook
    except ImportError:
        print("WARNING: openpyxl not installed; skipping Excel consolidation.")
        return False

    main_rows_by_size = []
    reduced_rows_by_size = []
    for size_name, _script in selected_sizes:
        main_dir = repo_root / "results" / size_name
        red_dir = main_dir / "dcm_9ants"
        expected = _expected_instances(repo_root, size_name)

        main_acs = _aggregate_alg_over_runs(
            outdir=main_dir, size_name=size_name, alg_name="CP-ACS",
            run_start=run_start, run_end=run_end, expected_instance_names=expected)
        main_dcm = _aggregate_alg_over_runs(
            outdir=main_dir, size_name=size_name, alg_name="CP-DCM-ACO",
            run_start=run_start, run_end=run_end, expected_instance_names=expected)
        red_dcm = _aggregate_alg_over_runs(
            outdir=red_dir, size_name=size_name, alg_name="CP-DCM-ACO",
            run_start=run_start, run_end=run_end, expected_instance_names=expected)
        for label, metrics in [
            (f"{size_name} CP-ACS", main_acs),
            (f"{size_name} CP-DCM-ACO", main_dcm),
            (f"{size_name} CP-DCM-ACO (9 ants)", red_dcm),
        ]:
            incomplete = int(metrics.get("runs_incomplete", 0))
            if incomplete > 0:
                print(
                    f"[{_now()}] NOTE: {label} has {incomplete} incomplete run(s); "
                    f"values use available rows in those runs.",
                    flush=True,
                )

        main_rows_by_size.append((size_name, [
            [
                "CP-ACS",
                _fmt(main_acs["best_success"], 3),
                _fmt(main_acs["worst_success"], 3),
                _fmt(main_acs["avg_success"], 3),
                _fmt(main_acs["time_mean"], 6),
                _fmt(main_acs["time_std"], 6),
                _fmt(main_acs["cycles_mean"], 3),
            ],
            [
                "CP-DCM-ACO",
                _fmt(main_dcm["best_success"], 3),
                _fmt(main_dcm["worst_success"], 3),
                _fmt(main_dcm["avg_success"], 3),
                _fmt(main_dcm["time_mean"], 6),
                _fmt(main_dcm["time_std"], 6),
                _fmt(main_dcm["cycles_mean"], 3),
            ],
        ]))

        reduced_rows_by_size.append((size_name, [
            [
                "CP-ACS",
                _fmt(main_acs["best_success"], 3),
                _fmt(main_acs["worst_success"], 3),
                _fmt(main_acs["avg_success"], 3),
                _fmt(main_acs["time_mean"], 6),
                _fmt(main_acs["time_std"], 6),
                _fmt(main_acs["cycles_mean"], 3),
            ],
            [
                "CP-DCM-ACO (9 ants)",
                _fmt(red_dcm["best_success"], 3),
                _fmt(red_dcm["worst_success"], 3),
                _fmt(red_dcm["avg_success"], 3),
                _fmt(red_dcm["time_mean"], 6),
                _fmt(red_dcm["time_std"], 6),
                _fmt(red_dcm["cycles_mean"], 3),
            ],
        ]))

    wb = Workbook()
    ws_main = wb.active
    ws_main.title = "main experiment"
    _write_sheet_tables(ws_main, "Main Experiment Summary", main_rows_by_size)

    ws_red = wb.create_sheet(title="reduced ant")
    _write_sheet_tables(ws_red, "Reduced Ant Summary", reduced_rows_by_size)

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(excel_path))
    print(f"[{_now()}] Consolidated Excel saved: {excel_path}", flush=True)
    return True


def main() -> int:
    repo_root = Path(__file__).resolve().parents[1]
    default_python = sys.executable

    ap = argparse.ArgumentParser(
        description=(
            "Run CP-ACS (alg 0) and CP-DCM-ACO (alg 2) with resume support, "
            "size-first ordering, and configurable repetitions."
        )
    )
    ap.add_argument("--run-start", type=int, default=1, help="First run index (default: 1)")
    ap.add_argument("--run-end", type=int, default=5, help="Last run index inclusive (default: 5)")
    ap.add_argument(
        "--reps",
        type=int,
        default=1,
        help="Repetitions per instance passed to child scripts (default: 1)",
    )
    ap.add_argument(
        "--pool-workers",
        type=int,
        default=4,
        help="Pool workers per algorithm process (default: 4)",
    )
    ap.add_argument(
        "--phase2-pool-workers",
        type=int,
        default=8,
        help="Pool workers for Phase 2 (DCM-ACO 9-ants). Default: 8",
    )
    ap.add_argument(
        "--binary",
        default=None,
        help="Optional explicit solver binary path passed to child scripts",
    )
    ap.add_argument(
        "--python",
        default=default_python,
        help="Python executable for child script invocations (default: current interpreter)",
    )
    ap.add_argument(
        "--dry-run",
        action="store_true",
        help="Print planned commands only; do not execute child processes",
    )
    ap.add_argument(
        "--size",
        action="append",
        choices=sorted(SIZE_ALIASES.keys()),
        help="Limit execution to one or more sizes: 9|16|25|9x9|16x16|25x25",
    )
    ap.add_argument(
        "--alg",
        action="append",
        type=int,
        choices=[0, 2],
        help="Limit execution to one or more algorithms (0 or 2). Default: both",
    )
    ap.add_argument(
        "--disable-transfer",
        action="store_true",
        help=(
            "Disable cross-alg worker transfer. By default, when both algs are selected "
            "and one finishes first, the other is restarted with additional pool workers."
        ),
    )
    ap.add_argument(
        "--skip-dcm9ants-phase",
        action="store_true",
        help="Skip the extra DCM-ACO 9-ants phase after CP comparison runs",
    )
    ap.add_argument(
        "--excel-path",
        default="results/cp_comparison_summary.xlsx",
        help="Path to consolidated Excel output (default: results/cp_comparison_summary.xlsx)",
    )
    ap.add_argument(
        "--no-consolidate",
        action="store_true",
        help="Skip Excel consolidation at the end",
    )
    ap.add_argument(
        "--consolidate-only",
        action="store_true",
        help="Build Excel from existing CSV results only (no benchmark runs)",
    )
    ap.add_argument("--verbose", action="store_true", help="Pass --verbose to child scripts")
    args = ap.parse_args()

    if args.run_start < 1:
        ap.error("--run-start must be >= 1")
    if args.run_end < args.run_start:
        ap.error("--run-end must be >= --run-start")
    if args.reps < 1:
        ap.error("--reps must be >= 1")
    if args.pool_workers < 1:
        ap.error("--pool-workers must be >= 1")
    if args.phase2_pool_workers < 1:
        ap.error("--phase2-pool-workers must be >= 1")

    if args.size:
        wanted = {SIZE_ALIASES[s] for s in args.size}
        selected_sizes = [(n, p) for (n, p) in SIZES if n in wanted]
    else:
        selected_sizes = list(SIZES)

    wanted_algs = set(args.alg) if args.alg else {0, 2}
    selected_algs = [(aid, aname) for (aid, aname) in ALGS if aid in wanted_algs]
    if not selected_algs:
        ap.error("No algorithms selected")
    excel_path = Path(args.excel_path)

    if args.consolidate_only:
        consolidate_excel(
            repo_root=repo_root,
            selected_sizes=selected_sizes,
            run_start=args.run_start,
            run_end=args.run_end,
            excel_path=excel_path,
        )
        return 0

    print(
        f"[{_now()}] Starting CP comparison reruns: "
        f"runs={args.run_start}..{args.run_end}, reps={args.reps}, "
        f"pool_workers={args.pool_workers}",
        flush=True,
    )

    # Phase 1: CP comparison (alg 0 + alg 2)
    for size_name, script_rel in selected_sizes:
        print(f"\n[{_now()}] ===== PHASE 1 | SIZE {size_name} START =====", flush=True)
        for run_idx in range(args.run_start, args.run_end + 1):
            print(
                f"[{_now()}] ----- PHASE 1 | SIZE {size_name} | RUN {run_idx} START -----",
                flush=True,
            )
            rc = _run_size_phase(
                repo_root=repo_root,
                python_exe=args.python,
                size_name=size_name,
                script_rel=script_rel,
                run_idx=run_idx,
                reps=args.reps,
                pool_workers=args.pool_workers,
                outdir=_default_outdir(size_name),
                extra_solver_args=None,
                binary=args.binary,
                algs=selected_algs,
                enable_transfer=(not args.disable_transfer),
                verbose=args.verbose,
                dry_run=args.dry_run,
            )
            if rc != 0:
                return rc
            print(
                f"[{_now()}] ----- PHASE 1 | SIZE {size_name} | RUN {run_idx} DONE -----",
                flush=True,
            )
        print(f"[{_now()}] ===== PHASE 1 | SIZE {size_name} DONE =====", flush=True)

    if args.skip_dcm9ants_phase:
        print(f"\n[{_now()}] Skipping DCM-ACO 9-ants phase by request.", flush=True)
        if not args.no_consolidate:
            consolidate_excel(
                repo_root=repo_root,
                selected_sizes=selected_sizes,
                run_start=args.run_start,
                run_end=args.run_end,
                excel_path=excel_path,
            )
        print(f"\n[{_now()}] All requested runs completed successfully.", flush=True)
        return 0

    # Phase 2: DCM-ACO (alg 2) with 9 total ants setup:
    # nAnts=3 per colony, numACS=2 => numColonies=3 (2 ACS + 1 MMAS)
    dcm9_args = ["--nAnts", "3", "--numACS", "2"]
    dcm9_algs = [(2, "CP-DCM-ACO-9ANTS")]
    for size_name, script_rel in selected_sizes:
        print(f"\n[{_now()}] ===== PHASE 2 (DCM-ACO 9 ants) | SIZE {size_name} START =====", flush=True)
        for run_idx in range(args.run_start, args.run_end + 1):
            print(
                f"[{_now()}] ----- PHASE 2 | SIZE {size_name} | RUN {run_idx} START -----",
                flush=True,
            )
            rc = _run_size_phase(
                repo_root=repo_root,
                python_exe=args.python,
                size_name=size_name,
                script_rel=script_rel,
                run_idx=run_idx,
                reps=1,
                pool_workers=args.phase2_pool_workers,
                outdir=_dcm9_outdir(size_name),
                extra_solver_args=dcm9_args,
                binary=args.binary,
                algs=dcm9_algs,
                enable_transfer=False,
                verbose=args.verbose,
                dry_run=args.dry_run,
            )
            if rc != 0:
                return rc
            print(
                f"[{_now()}] ----- PHASE 2 | SIZE {size_name} | RUN {run_idx} DONE -----",
                flush=True,
            )
        print(f"[{_now()}] ===== PHASE 2 | SIZE {size_name} DONE =====", flush=True)

    if not args.no_consolidate:
        consolidate_excel(
            repo_root=repo_root,
            selected_sizes=selected_sizes,
            run_start=args.run_start,
            run_end=args.run_end,
            excel_path=excel_path,
        )

    print(f"\n[{_now()}] All requested runs completed successfully.", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
